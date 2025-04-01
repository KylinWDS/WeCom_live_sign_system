from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
    QLabel, QLineEdit, QTableWidget, QTableWidgetItem,
    QMessageBox, QComboBox, QSpinBox, QHeaderView,
    QFileDialog, QDialog, QFormLayout, QGroupBox,
    QDateEdit, QTimeEdit, QToolBar, QSpacerItem, QSizePolicy,
    QProgressDialog, QInputDialog, QTextEdit, QGridLayout, QMenu,
    QStyleOption, QStyle, QApplication, QDoubleSpinBox
)
from PySide6.QtCore import Qt, QDateTime, QTimer, QPoint, QSortFilterProxyModel, QRect, QSize
from PySide6.QtGui import QIcon, QPainter, QColor, QPen, QBrush, QFontMetrics
from ..managers.style import StyleManager
from ..utils.widget_utils import WidgetUtils
from src.utils.logger import get_logger
from src.ui.managers.animation import AnimationManager
from src.utils.performance_manager import PerformanceManager
from src.utils.error_handler import ErrorHandler
from src.models.live_booking import LiveBooking, LiveStatus as BookingStatus
from src.core.database import DatabaseManager
from src.api.wecom import WeComAPI
from src.models.living import Living, LivingStatus, LivingType
from datetime import datetime, timedelta
from src.ui.components.dialogs.io_dialog import IODialog
from src.ui.components.widgets.custom_datetime_widget import CustomDateTimeWidget
from src.models.live_viewer import LiveViewer
from src.models.live_sign_record import LiveSignRecord
from src.models.user import User, UserRole
import pandas as pd
import os
from typing import List, Optional, Dict, Any
from sqlalchemy import select, func, and_, or_, cast, String, text
from src.core.live_viewer_manager import LiveViewerManager
import concurrent.futures
from threading import Lock
from copy import deepcopy
from collections import defaultdict, namedtuple
from src.models.live_reward_record import LiveRewardRecord, RewardRuleType
import logging

logger = get_logger(__name__)

class AutoCloseMessageBox(QMessageBox):
    """自动关闭的消息对话框"""
    
    def __init__(self, title, text, timeout=3000, parent=None):
        """
        初始化自动关闭的消息对话框
        
        Args:
            title: 标题
            text: 显示文本
            timeout: 自动关闭的时间（毫秒）
            parent: 父窗口
        """
        super().__init__(QMessageBox.Information, title, text, QMessageBox.NoButton, parent)
        self.timeout = timeout
        
        # 添加倒计时标签
        self.time_label = QLabel(f"{timeout//1000}秒后自动关闭", self)
        self.time_label.setStyleSheet("color: gray; margin-top: 10px;")
        self.time_label.setAlignment(Qt.AlignRight)
        
        # 获取布局并添加标签
        layout = self.layout()
        layout.addWidget(self.time_label, layout.rowCount(), 0, 1, layout.columnCount(), Qt.AlignRight)
        
        # 设置定时器
        self.countdown = timeout // 1000  # 转换为秒
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.update_counter)
        self.timer.start(1000)  # 每秒更新一次
    
    def update_counter(self):
        """更新倒计时"""
        self.countdown -= 1
        if self.countdown <= 0:
            self.timer.stop()
            self.close()
        else:
            self.time_label.setText(f"{self.countdown}秒后自动关闭")

class LiveListPage(QWidget):
    """直播列表页面"""
    
    def __init__(self, db_manager: DatabaseManager, wecom_api: WeComAPI, auth_manager=None, user_id=None):
        super().__init__()
        self.db_manager = db_manager
        self.wecom_api = wecom_api
        self.auth_manager = auth_manager
        self.user_id = user_id
        self.performance_manager = PerformanceManager()
        self.error_handler = ErrorHandler()
        self.init_ui()
        
    def init_ui(self):
        """初始化UI"""
        self.setObjectName("liveListPage")
        
        # 创建主布局
        layout = QVBoxLayout(self)
        layout.setSpacing(20)
        layout.setContentsMargins(0, 0, 0, 0)  # 减少外边距，使内容能够占满更多空间
        
        # 创建搜索和工具栏区域
        search_tool_layout = QVBoxLayout()
        search_tool_layout.setSpacing(10)
        search_tool_layout.setContentsMargins(0, 0, 0, 0)  # 减少边距
        
        # 创建搜索区域
        search_group = self._create_search_group()
        search_tool_layout.addWidget(search_group)
        
        # 添加工具栏
        toolbar = self._create_toolbar()
        search_tool_layout.addWidget(toolbar)
        
        layout.addLayout(search_tool_layout)
        
        # 创建表格
        self.table = QTableWidget()
        self.table.setColumnCount(17)  # 增加到17列，添加序号列
        self.table.setHorizontalHeaderLabels([
            "序号", "直播ID", "直播标题", "开始时间", "结束时间",
            "主播", "状态", "直播类型", 
            "观看人数", "评论数", "签到人数", "签到次数", 
            "观看信息", "签到导入", "企微文档", "远程同步", "操作"
        ])
        
        # 设置表格可以显示滚动条
        self.table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOn)  # 始终显示横向滚动条
        self.table.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        
        # 设置各列的宽度模式 - 除了操作列使用固定宽度外，其他列使用固定宽度以确保内容可见
        column_widths = [60, 100, 200, 150, 150, 150, 100, 120, 80, 80, 80, 80, 80, 80, 80, 80, 350]  # 添加序号列宽度
        
        for i in range(16):  # 前16列使用固定宽度
            self.table.horizontalHeader().setSectionResizeMode(i, QHeaderView.Interactive)  # 允许用户调整宽度
            self.table.setColumnWidth(i, column_widths[i])
            
        # 操作列使用固定宽度，确保按钮显示完整
        self.table.horizontalHeader().setSectionResizeMode(16, QHeaderView.Fixed)
        self.table.setColumnWidth(16, column_widths[16])  # 为操作列设置足够宽度
        
        # 设置垂直头部策略，适应内容
        self.table.verticalHeader().setVisible(False)  # 隐藏行号
        self.table.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # 设置选择模式
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        # 应用样式
        WidgetUtils.set_table_style(self.table)
        layout.addWidget(self.table)
        
        # 创建分页区域
        pagination_layout = QHBoxLayout()
        
        # 上一页按钮
        self.prev_btn = QPushButton("上一页")
        self.prev_btn.setObjectName("secondaryButton")
        self.prev_btn.clicked.connect(self.prev_page)
        pagination_layout.addWidget(self.prev_btn)
        
        # 页码显示
        self.page_label = QLabel("第 1 页")
        pagination_layout.addWidget(self.page_label)
        
        # 下一页按钮
        self.next_btn = QPushButton("下一页")
        self.next_btn.setObjectName("secondaryButton")
        self.next_btn.clicked.connect(self.next_page)
        pagination_layout.addWidget(self.next_btn)
        
        layout.addLayout(pagination_layout)
        
        # 设置样式
        self.setStyleSheet(StyleManager.get_live_list_style())
        
        # 添加淡入动画
        AnimationManager.fade_in(self)
        
        # 加载数据
        self.current_page = 1
        self.page_size = 10
        self.load_data()
        
    def _create_search_group(self) -> QGroupBox:
        """创建搜索区域"""
        group = QGroupBox("搜索条件")
        
        # 设置尺寸策略，使搜索区域占满整个屏幕宽度
        group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        main_layout = QVBoxLayout(group)
        main_layout.setSpacing(10)
        main_layout.setContentsMargins(10, 15, 10, 15)  # 设置适当的内边距
        
        # 第一行：标题和状态
        first_row = QHBoxLayout()
        first_row.setSpacing(20)  # 增加间距，改善视觉效果
        
        # 直播标题
        title_label = QLabel("直播标题:")
        title_label.setMinimumWidth(60)  # 设置最小宽度使标签对齐
        first_row.addWidget(title_label)
        
        self.live_title = QLineEdit()
        # 移除最大宽度限制，使用策略让输入框伸展
        self.live_title.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        WidgetUtils.set_input_style(self.live_title)
        first_row.addWidget(self.live_title, 3)  # 给标题输入框分配更多空间
        
        # 添加一些间距
        first_row.addSpacing(20)
        
        # 直播状态
        status_label = QLabel("直播状态:")
        status_label.setMinimumWidth(60)  # 设置最小宽度使标签对齐
        first_row.addWidget(status_label)
        
        self.live_status = QComboBox()
        self.live_status.addItems(["全部", "未开始", "进行中", "已结束"])
        # 设置合适的宽度但不限制最大宽度
        self.live_status.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.live_status.setMinimumWidth(100)  # 确保下拉框有足够宽度
        WidgetUtils.set_combo_style(self.live_status)
        first_row.addWidget(self.live_status, 1)
        
        # 添加弹性空间，让元素分布均匀
        first_row.addStretch(1)
        
        # 第二行：新增的状态字段搜索
        status_row = QHBoxLayout()
        status_row.setSpacing(20)  # 增加间距
        
        # 为每个状态字段设置固定的标签宽度，使其对齐
        label_width = 70
        
        # 是否拉取观看信息
        viewer_label = QLabel("观看信息:")
        viewer_label.setMinimumWidth(label_width)
        status_row.addWidget(viewer_label)
        
        self.viewer_fetched_status = QComboBox()
        self.viewer_fetched_status.addItems(["全部", "已拉取", "未拉取"])
        # 确保下拉框有合适的尺寸
        self.viewer_fetched_status.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.viewer_fetched_status.setMinimumWidth(80)
        WidgetUtils.set_combo_style(self.viewer_fetched_status)
        status_row.addWidget(self.viewer_fetched_status)
        
        # 添加一些间距
        status_row.addSpacing(10)
        
        # 是否导入签到
        sign_label = QLabel("签到导入:")
        sign_label.setMinimumWidth(label_width)
        status_row.addWidget(sign_label)
        
        self.sign_imported_status = QComboBox()
        self.sign_imported_status.addItems(["全部", "已导入", "未导入"])
        self.sign_imported_status.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.sign_imported_status.setMinimumWidth(80)
        WidgetUtils.set_combo_style(self.sign_imported_status)
        status_row.addWidget(self.sign_imported_status)
        
        # 添加一些间距
        status_row.addSpacing(10)
        
        # 是否上传企微文档
        doc_label = QLabel("企微文档:")
        doc_label.setMinimumWidth(label_width)
        status_row.addWidget(doc_label)
        
        self.doc_uploaded_status = QComboBox()
        self.doc_uploaded_status.addItems(["全部", "已上传", "未上传"])
        self.doc_uploaded_status.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.doc_uploaded_status.setMinimumWidth(80)
        WidgetUtils.set_combo_style(self.doc_uploaded_status)
        status_row.addWidget(self.doc_uploaded_status)
        
        # 添加一些间距
        status_row.addSpacing(10)
        
        # 是否远程同步
        remote_label = QLabel("远程同步:")
        remote_label.setMinimumWidth(label_width)
        status_row.addWidget(remote_label)
        
        self.remote_synced_status = QComboBox()
        self.remote_synced_status.addItems(["全部", "已同步", "未同步"])
        self.remote_synced_status.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Fixed)
        self.remote_synced_status.setMinimumWidth(80)
        WidgetUtils.set_combo_style(self.remote_synced_status)
        status_row.addWidget(self.remote_synced_status)
        
        # 添加弹性空间，让元素分布均匀
        status_row.addStretch(1)
        
        # 日期时间范围行
        date_row = QHBoxLayout()
        date_row.setSpacing(20)  # 增加间距
        
        # 直播时间范围 - 使用自定义日期时间组件
        start_label = QLabel("开始时间:")
        start_label.setMinimumWidth(60)
        date_row.addWidget(start_label)
        
        self.start_date_time = CustomDateTimeWidget()
        self.start_date_time.setDateTime(QDateTime.currentDateTime().addMonths(-1))  # 默认一个月前
        # 使日期时间控件能够适应可用空间
        self.start_date_time.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        date_row.addWidget(self.start_date_time)
        
        # 添加"至"标签，并设置固定宽度
        to_label = QLabel("至")
        to_label.setFixedWidth(20)
        to_label.setAlignment(Qt.AlignCenter)
        date_row.addWidget(to_label)
        
        self.end_date_time = CustomDateTimeWidget()
        self.end_date_time.setDateTime(QDateTime.currentDateTime())  # 默认今天
        # 使日期时间控件能够适应可用空间
        self.end_date_time.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        date_row.addWidget(self.end_date_time)
        
        # 添加弹性空间，让元素分布均匀
        date_row.addStretch(1)
        
        # 第三行：按钮行
        button_row = QHBoxLayout()
        button_row.setSpacing(20)  # 增加间距
        
        # 在按钮行左侧添加弹性空间
        button_row.addStretch(1)
        
        # 搜索按钮
        search_btn = QPushButton("搜索")
        search_btn.setObjectName("primaryButton")
        search_btn.setMinimumWidth(100)  # 设置最小宽度，使按钮看起来更均匀
        search_btn.clicked.connect(self.search)
        button_row.addWidget(search_btn)
        
        # 重置按钮
        reset_btn = QPushButton("重置")
        reset_btn.setObjectName("primaryButton")
        reset_btn.setMinimumWidth(100)
        reset_btn.clicked.connect(self.reset_search)
        button_row.addWidget(reset_btn)
        
        # 同步按钮
        sync_btn = QPushButton("同步直播数据")
        sync_btn.setObjectName("primaryButton")
        sync_btn.setMinimumWidth(120)
        sync_btn.clicked.connect(self.sync_live_data)
        button_row.addWidget(sync_btn)
        
        # 导出按钮
        export_btn = QPushButton("导出数据")
        export_btn.setObjectName("primaryButton")
        export_btn.setMinimumWidth(100)
        export_btn.clicked.connect(self.export_data)
        button_row.addWidget(export_btn)
        
        # 综合导出数据按钮
        combined_export_btn = QPushButton("综合导出数据")
        combined_export_btn.setObjectName("primaryButton")
        combined_export_btn.setMinimumWidth(120)
        combined_export_btn.clicked.connect(self.show_combined_export_dialog)
        button_row.addWidget(combined_export_btn)
        
        # 在按钮行右侧添加弹性空间
        button_row.addStretch(1)
        
        # 将三行添加到主布局
        main_layout.addLayout(first_row)
        main_layout.addLayout(status_row)
        main_layout.addLayout(date_row)
        main_layout.addLayout(button_row)
        
        return group
        
    @PerformanceManager.measure_operation("load_data")
    def load_data(self):
        """加载数据"""
        try:
            # 获取直播列表
            records_data = []  # 存储转换后的记录数据
            
            with self.db_manager.get_session() as session:
                # 获取当前用户信息和权限
                current_user = None
                user_role = None
                user_corpname = None
                user_id = None
                
                if self.auth_manager and self.user_id:
                    # 在当前会话中获取用户信息
                    from src.models.user import User, UserRole
                    current_user = session.query(User).filter_by(userid=self.user_id).first()
                    if current_user:
                        user_role = current_user.role
                        user_corpname = current_user.corpname
                        user_id = current_user.wecom_code or current_user.login_name
                
                # 使用旧版API风格
                query = session.query(Living)
                
                # 根据用户权限过滤数据
                if current_user:
                    if user_role == UserRole.ROOT_ADMIN.value:
                        # 超级管理员可以查看所有直播
                        pass
                    elif user_role == UserRole.WECOM_ADMIN.value and user_corpname:
                        # 企业管理员只能查看自己企业的直播
                        query = query.filter(Living.corpname == user_corpname)
                    else:
                        # 普通用户只能查看自己为主播的直播
                        if user_id:
                            query = query.filter(Living.anchor_userid == user_id)
                        else:
                            # 如果没有企业微信ID，则显示空列表
                            logger.warning(f"用户 {current_user.login_name} 没有企业微信ID，无法显示直播列表")
                            self.table.setRowCount(0)
                            self.prev_btn.setEnabled(False)
                            self.next_btn.setEnabled(False)
                            self.page_label.setText("第 0 页 / 共 0 页")
                            return
                
                # 应用搜索条件
                if self.live_title.text():
                    query = query.filter(Living.theme.like(f"%{self.live_title.text()}%"))
                    
                if self.live_status.currentText() != "全部":
                    status_map = {
                        "未开始": LivingStatus.RESERVED,
                        "进行中": LivingStatus.LIVING,
                        "已结束": LivingStatus.ENDED
                    }
                    query = query.filter(Living.status == status_map[self.live_status.currentText()])
                
                # 应用新增的状态字段过滤条件
                # 是否拉取观看信息
                if self.viewer_fetched_status.currentText() != "全部":
                    is_fetched = 1 if self.viewer_fetched_status.currentText() == "已拉取" else 0
                    query = query.filter(Living.is_viewer_fetched == is_fetched)
                
                # 是否导入签到
                if self.sign_imported_status.currentText() != "全部":
                    is_imported = 1 if self.sign_imported_status.currentText() == "已导入" else 0
                    query = query.filter(Living.is_sign_imported == is_imported)
                
                # 是否上传企微文档
                if self.doc_uploaded_status.currentText() != "全部":
                    is_uploaded = 1 if self.doc_uploaded_status.currentText() == "已上传" else 0
                    query = query.filter(Living.is_doc_uploaded == is_uploaded)
                
                # 是否远程同步
                if self.remote_synced_status.currentText() != "全部":
                    is_synced = 1 if self.remote_synced_status.currentText() == "已同步" else 0
                    query = query.filter(Living.is_remote_synced == is_synced)
                
                # 应用日期范围查询
                start_datetime = self.start_date_time.dateTime().toPython()
                end_datetime = self.end_date_time.dateTime().toPython()
                
                # 过滤直播开始时间在指定范围内的记录
                query = query.filter(Living.living_start >= start_datetime)
                query = query.filter(Living.living_start <= end_datetime)
                
                # 计算总记录数
                total = query.count()
                self.total_pages = (total + self.page_size - 1) // self.page_size
                
                # 获取当前页数据
                records = query.offset((self.current_page - 1) * self.page_size).limit(self.page_size).all()
                
                # 在会话内将数据转换为字典，避免会话关闭后的访问问题
                for record in records:
                    # 获取所有需要的数据
                    record_data = {
                        "id": record.id,  # 用于标识记录
                        "livingid": record.livingid,
                        "theme": record.theme,
                        "living_start": record.living_start,
                        "living_duration": record.living_duration,
                        "anchor_userid": record.anchor_userid,
                        "description": record.description,
                        "type": record.type,
                        "status": record.status,
                        "corpname": record.corpname,
                        "agentid": record.agentid,
                        "viewer_num": record.viewer_num,
                        "comment_num": record.comment_num,
                        "mic_num": record.mic_num,
                        "online_count": record.online_count,
                        "subscribe_count": record.subscribe_count
                    }
                    
                    # 获取主播的名称
                    from src.models.user import User
                    user = session.query(User).filter(
                        (User.wecom_code == record.anchor_userid) | 
                        (User.login_name == record.anchor_userid)
                    ).first()
                    
                    if user:
                        record_data["anchor_name"] = f"{user.name}({record.anchor_userid})"
                    else:
                        record_data["anchor_name"] = f"{record.anchor_userid}"
                    
                    # 计算结束时间
                    if record.living_start and record.living_duration:
                        record_data["end_time"] = record.living_start + timedelta(seconds=record.living_duration)
                    else:
                        record_data["end_time"] = None
                        
                    # 获取签到统计信息
                    from src.models.live_viewer import LiveViewer
                    # 使用 LiveViewer 查询签到统计
                    sign_stats = {}
                    with self.db_manager.get_session() as stats_session:
                        # 获取不同的签到人数
                        unique_signers = stats_session.query(func.count(LiveViewer.id)).filter(
                            LiveViewer.living_id == record.id,
                            LiveViewer.is_signed == True
                        ).scalar() or 0
                        
                        # 获取总签到次数
                        sign_count = stats_session.query(func.sum(LiveViewer.sign_count)).filter(
                            LiveViewer.living_id == record.id,
                            LiveViewer.is_signed == True
                        ).scalar() or 0
                        
                        # 获取首次签到时间
                        first_sign = stats_session.query(func.min(LiveViewer.sign_time)).filter(
                            LiveViewer.living_id == record.id,
                            LiveViewer.is_signed == True
                        ).scalar()
                        
                        sign_stats = {
                            "unique_signers": unique_signers,
                            "sign_count": sign_count,
                            "sign_time": first_sign
                        }
                    
                    record_data["sign_count"] = sign_stats["unique_signers"]  # 使用unique_signers作为签到人数
                    record_data["total_sign_count"] = sign_stats["sign_count"]  # 总签到次数
                    record_data["sign_time"] = sign_stats["sign_time"]
                    
                    # 添加状态字段
                    record_data["is_viewer_fetched"] = record.is_viewer_fetched
                    record_data["is_sign_imported"] = record.is_sign_imported
                    record_data["is_doc_uploaded"] = record.is_doc_uploaded
                    record_data["is_remote_synced"] = record.is_remote_synced
                    
                    records_data.append(record_data)
                
            # 更新表格
            self.table.setRowCount(len(records_data))
            for row, record_data in enumerate(records_data):
                # 设置足够的行高以容纳按钮
                self.table.setRowHeight(row, 27)  # 设置每行高度为原来的2/3（约27像素）
                
                # 设置序号
                self.table.setItem(row, 0, QTableWidgetItem(str(row + 1 + (self.current_page - 1) * self.page_size)))
                
                # 设置直播ID
                self.table.setItem(row, 1, QTableWidgetItem(record_data["livingid"]))
                self.table.setItem(row, 2, QTableWidgetItem(record_data["theme"]))
                self.table.setItem(row, 3, QTableWidgetItem(record_data["living_start"].strftime("%Y-%m-%d %H:%M:%S")))
                
                # 结束时间
                end_time_str = record_data["end_time"].strftime("%Y-%m-%d %H:%M:%S") if record_data["end_time"] else "-"
                self.table.setItem(row, 4, QTableWidgetItem(end_time_str))
                
                # 主播名称
                self.table.setItem(row, 5, QTableWidgetItem(record_data["anchor_name"]))
                
                # 状态
                status_text = {
                    LivingStatus.RESERVED: "预约中",
                    LivingStatus.LIVING: "直播中",
                    LivingStatus.ENDED: "已结束",
                    LivingStatus.EXPIRED: "已过期",
                    LivingStatus.CANCELLED: "已取消"
                }.get(record_data["status"], "未知")
                self.table.setItem(row, 6, QTableWidgetItem(status_text))
                
                # 直播类型
                type_text = {
                    LivingType.GENERAL: "通用直播",
                    LivingType.SMALL: "小班课",
                    LivingType.LARGE: "大班课",
                    LivingType.TRAINING: "企业培训",
                    LivingType.EVENT: "活动直播"
                }.get(record_data["type"], "未知")
                self.table.setItem(row, 7, QTableWidgetItem(type_text))
                
                self.table.setItem(row, 8, QTableWidgetItem(str(record_data["viewer_num"])))
                self.table.setItem(row, 9, QTableWidgetItem(str(record_data["comment_num"])))
                
                # 签到人数
                self.table.setItem(row, 10, QTableWidgetItem(str(record_data["sign_count"])))
                
                # 签到次数
                self.table.setItem(row, 11, QTableWidgetItem(str(record_data["total_sign_count"])))
                
                # 观看信息状态
                viewer_fetched_text = "已拉取" if record_data["is_viewer_fetched"] == 1 else "未拉取"
                viewer_item = QTableWidgetItem(viewer_fetched_text)
                viewer_item.setForeground(Qt.green if record_data["is_viewer_fetched"] == 1 else Qt.red)
                self.table.setItem(row, 12, viewer_item)
                
                # 签到导入状态
                sign_imported_text = "已导入" if record_data["is_sign_imported"] == 1 else "未导入"
                sign_item = QTableWidgetItem(sign_imported_text)
                sign_item.setForeground(Qt.green if record_data["is_sign_imported"] == 1 else Qt.red)
                self.table.setItem(row, 13, sign_item)
                
                # 企微文档状态
                doc_uploaded_text = "已上传" if record_data["is_doc_uploaded"] == 1 else "未上传"
                doc_item = QTableWidgetItem(doc_uploaded_text)
                doc_item.setForeground(Qt.green if record_data["is_doc_uploaded"] == 1 else Qt.red)
                self.table.setItem(row, 14, doc_item)
                
                # 远程同步状态
                remote_synced_text = "已同步" if record_data["is_remote_synced"] == 1 else "未同步"
                remote_item = QTableWidgetItem(remote_synced_text)
                remote_item.setForeground(Qt.green if record_data["is_remote_synced"] == 1 else Qt.red)
                self.table.setItem(row, 15, remote_item)
                
                # 操作按钮
                btn_widget = QWidget()
                btn_layout = QHBoxLayout(btn_widget)
                btn_layout.setContentsMargins(2, 2, 2, 2)  # 减小边距使按钮更紧凑
                btn_layout.setSpacing(5)  # 减少按钮之间的间距
                
                # 创建Living对象用于按钮回调，避免会话问题
                living = Living(
                    livingid=record_data["livingid"],
                    theme=record_data["theme"],
                    living_start=record_data["living_start"],
                    living_duration=record_data["living_duration"],
                    anchor_userid=record_data["anchor_userid"],
                    description=record_data["description"],
                    status=record_data["status"],
                    type=record_data["type"],
                    corpname=record_data["corpname"],
                    agentid=record_data["agentid"],
                    viewer_num=record_data["viewer_num"],
                    comment_num=record_data["comment_num"],
                    mic_num=record_data["mic_num"],
                    online_count=record_data["online_count"],
                    subscribe_count=record_data["subscribe_count"]
                )
                # 保存签到信息到living对象中，以便在详情对话框中使用
                living.sign_count = record_data["sign_count"]
                living.total_sign_count = record_data["total_sign_count"]
                living.sign_time = record_data["sign_time"]
                
                # 创建样式化的按钮，减小高度以适应行高
                view_btn = QPushButton("查看详情")
                view_btn.setObjectName("linkButton")
                view_btn.setMinimumHeight(22)  # 减小按钮高度
                view_btn.setFixedWidth(80)  # 固定宽度
                view_btn.clicked.connect(lambda checked, r=living: self.view_details(r))
                btn_layout.addWidget(view_btn)
                
                # 添加"拉取观看信息"按钮，并根据状态调整显示
                fetch_viewer_btn = QPushButton("拉取观看信息" if not record_data["is_viewer_fetched"] else "重取观看信息")
                fetch_viewer_btn.setObjectName("linkButton")
                fetch_viewer_btn.setMinimumHeight(22)  # 减小按钮高度
                fetch_viewer_btn.setFixedWidth(120)  # 固定宽度
                # 如果已经拉取过，使用不同的样式
                if record_data["is_viewer_fetched"]:
                    fetch_viewer_btn.setStyleSheet("color: #0056b3;")  # 使用较暗的蓝色
                fetch_viewer_btn.clicked.connect(lambda checked, r=living: self.fetch_watch_stat(r))
                btn_layout.addWidget(fetch_viewer_btn)
                
                import_btn = QPushButton("导入签到")
                import_btn.setObjectName("linkButton")
                import_btn.setMinimumHeight(22)  # 减小按钮高度
                import_btn.setFixedWidth(80)  # 固定宽度
                import_btn.clicked.connect(lambda checked, r=living: self.import_sign(r))
                btn_layout.addWidget(import_btn)
                
                # 只有在直播状态为"预约中"时才显示取消按钮
                if record_data["status"] == LivingStatus.RESERVED:
                    cancel_btn = QPushButton("取消")
                    cancel_btn.setObjectName("linkButton")
                    cancel_btn.setMinimumHeight(22)  # 减小按钮高度
                    cancel_btn.setFixedWidth(80)  # 固定宽度
                    cancel_btn.clicked.connect(lambda checked, r=living: self.cancel_live(r))
                    btn_layout.addWidget(cancel_btn)
                
                self.table.setCellWidget(row, 16, btn_widget)
                
            # 更新分页信息
            self.page_label.setText(f"第 {self.current_page} 页 / 共 {self.total_pages} 页")
            self.prev_btn.setEnabled(self.current_page > 1)
            self.next_btn.setEnabled(self.current_page < self.total_pages)
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, "加载直播列表失败")
            
    def search(self):
        """搜索"""
        self.current_page = 1
        self.load_data()
        
    def prev_page(self):
        """上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_data()
            
    def next_page(self):
        """下一页"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_data()
            
    @PerformanceManager.measure_operation("view_details")
    def view_details(self, live: Living):
        """查看直播详情
        
        Args:
            live: 直播信息
        """
        try:
            # 从数据库获取最新记录，确保状态是最新的
            with self.db_manager.get_session() as session:
                record = session.query(Living).filter_by(livingid=live.livingid).first()
                if record:
                    # 使用最新的数据库记录创建详情对话框
                    dialog = LiveDetailsDialog(record)
                else:
                    # 如果找不到记录，使用传入的live对象
                    dialog = LiveDetailsDialog(live)
                dialog.exec()
                
            # 获取直播详情
            response = self.wecom_api.get_living_info(live.livingid)
            
            if response.get("errcode") == 0:
                # 获取直播信息
                live_info = response.get("living_info", {})
                
                # 更新直播数据
                with self.db_manager.get_session() as session:
                    # 从数据库中获取最新的实体
                    stmt = select(Living).where(Living.livingid == live.livingid)
                    result = session.execute(stmt)
                    db_live = result.scalar_one_or_none()
                    
                    # 如果找到实体，才更新它
                    if db_live:
                        db_live.viewer_num = live_info["viewer_num"]
                        db_live.comment_num = live_info["comment_num"]
                        db_live.mic_num = live_info["mic_num"]
                        
                        # 处理状态字段，确保它是 LivingStatus 枚举类型
                        status_value = live_info.get("status")
                        if status_value is not None:
                            if isinstance(status_value, int) or (isinstance(status_value, str) and status_value.isdigit()):
                                status_value = int(status_value)
                                # 将数字映射到 LivingStatus 枚举
                                status_map = {
                                    0: LivingStatus.RESERVED,
                                    1: LivingStatus.LIVING,
                                    2: LivingStatus.ENDED,
                                    3: LivingStatus.EXPIRED,
                                    4: LivingStatus.CANCELLED
                                }
                                db_live.status = status_map.get(status_value, LivingStatus.RESERVED)
            else:
                ErrorHandler.handle_warning(
                    f"获取直播详情失败：{response.get('errmsg')}",
                    self,
                    "失败"
                )
        except Exception as e:
            ErrorHandler.handle_error(e, self, "查看直播详情失败")
            
    @PerformanceManager.measure_operation("sync_live_data")
    def sync_live_data(self):
        """同步直播数据"""
        try:
            confirm = ErrorHandler.handle_question(
                "确定要同步直播数据吗？这将从企业微信拉取最新数据并更新本地数据库。",
                self,
                "确认同步"
            )
            
            if not confirm:
                return
                
            # 显示倒计时提示（自动3秒关闭）
            msg_box = AutoCloseMessageBox("同步进行中", "正在同步直播数据，请稍候...", 3000, self)
            msg_box.exec()
            
            # 获取用于API调用的用户ID列表
            user_ids_for_api = []
            
            with self.db_manager.get_session() as session:
                from src.models.user import UserRole, User
                
                # 1. 获取当前用户信息和权限
                current_user = None
                user_role = None
                user_corpid = None
                
                if self.auth_manager and self.user_id:
                    # 在当前会话中获取用户信息
                    current_user = session.query(User).filter_by(userid=self.user_id).first()
                    if current_user:
                        user_role = current_user.role
                        user_corpid = current_user.corpid
                
                if not current_user:
                    ErrorHandler.handle_warning(
                        "无法获取当前用户信息，将使用默认权限同步数据。",
                        self,
                        "警告"
                    )
                
                # 2. 根据权限确定要获取的用户ID列表
                if current_user and user_role == UserRole.ROOT_ADMIN.value:
                    # 如果是超级管理员，获取所有用户的直播列表
                    users = session.query(User).all()
                    for user in users:
                        userid = user.wecom_code or user.login_name
                        if userid:
                            user_ids_for_api.append(userid)
                
                elif current_user and user_role == UserRole.WECOM_ADMIN.value and user_corpid:
                    # 如果是企业管理员，获取该企业下所有用户的直播列表
                    users = session.query(User).filter_by(corpid=user_corpid).all()
                    for user in users:
                        userid = user.wecom_code or user.login_name
                        if userid:
                            user_ids_for_api.append(userid)
                
                else:
                    # 如果是普通用户或无法确定角色，只获取自己的直播列表
                    if current_user:
                        userid = current_user.wecom_code or current_user.login_name
                        if userid:
                            user_ids_for_api.append(userid)
            
            # 从企业微信API获取直播ID列表
            livingid_list = []
            # 创建一个映射，记录每个直播ID是从哪个用户获取的
            livingid_user_map = {}
            
            for userid in user_ids_for_api:
                try:
                    response = self.wecom_api.get_user_all_livingid(userid)
                    if response.get("errcode") == 0:
                        live_ids = response.get("livingid_list", [])
                        for live_id in live_ids:
                            # 记录这个直播ID是从哪个用户获取的
                            livingid_user_map[live_id] = userid
                        livingid_list.extend(live_ids)
                except Exception as e:
                    logger.error(f"获取用户 {userid} 的直播列表失败: {str(e)}")
            
            # 去重
            livingid_list = list(set(livingid_list))
            
            # 3. 从预约直播表中获取数据
            bookings = {}
            with self.db_manager.get_session() as session:
                booking_records = session.query(LiveBooking).all()
                for booking in booking_records:
                    bookings[booking.livingid] = {
                        "id": booking.id,  # 添加booking的ID
                        "livingid": booking.livingid,
                        "theme": booking.theme,
                        "living_start": booking.living_start,
                        "living_duration": booking.living_duration,
                        "anchor_userid": booking.anchor_userid,
                        "description": booking.description,
                        "type": booking.type,
                        "status": booking.status,
                        "corpname": booking.corpname,
                        "agentid": booking.agentid,
                        "viewer_num": booking.viewer_num,
                        "comment_num": booking.comment_num,
                        "mic_num": booking.mic_num,
                        "online_count": booking.online_count,
                        "subscribe_count": booking.subscribe_count,
                    }
            
            # 4. 从直播信息表中获取数据
            livings = {}
            with self.db_manager.get_session() as session:
                living_records = session.query(Living).all()
                for living in living_records:
                    livings[living.livingid] = {
                        "id": living.id,  # 添加living的ID
                        "livingid": living.livingid,
                        "theme": living.theme,
                        "living_start": living.living_start,
                        "living_duration": living.living_duration,
                        "anchor_userid": living.anchor_userid,
                        "description": living.description,
                        "type": living.type,
                        "status": living.status,
                        "corpname": living.corpname,
                        "agentid": living.agentid,
                        "viewer_num": living.viewer_num,
                        "comment_num": living.comment_num,
                        "mic_num": living.mic_num,
                        "online_count": living.online_count,
                        "subscribe_count": living.subscribe_count,
                    }
            
            # 5. 创建所有直播ID的集合（API获取的 + 本地数据库的）
            all_live_ids = set(livingid_list) | set(bookings.keys()) | set(livings.keys())
            
            # 6. 从企业微信获取直播数据
            updated_count = 0
            created_count = 0
            
            with self.db_manager.get_session() as session:
                # 创建用户ID到用户对象的映射，用于获取企业信息
                user_objects = {}
                from sqlalchemy import or_
                for userid in user_ids_for_api:
                    user = session.query(User).filter(
                        or_(User.wecom_code == userid, User.login_name == userid)
                    ).first()
                    if user:
                        user_objects[userid] = user
                
                for livingid in all_live_ids:
                    # 从企业微信获取数据
                    try:
                        response = self.wecom_api.get_living_info(livingid)
                        
                        if response.get("errcode") == 0:
                            # 获取企业微信返回的数据
                            wecom_data = response.get("living_info", {})
                            
                            # 创建最终数据（优先级：企业微信 > livings > bookings）
                            final_data = {}
                            
                            # 添加bookings数据
                            booking_id = None
                            if livingid in bookings:
                                final_data.update(bookings[livingid])
                                booking_id = bookings[livingid]["id"]
                                
                            # 添加livings数据
                            if livingid in livings:
                                final_data.update(livings[livingid])
                                
                            # 添加企业微信数据（转换格式）
                            if wecom_data:
                                # 转换时间和状态
                                if "living_start" in wecom_data and isinstance(wecom_data["living_start"], (int, float)):
                                    wecom_data["living_start"] = datetime.fromtimestamp(wecom_data["living_start"])
                                    
                                # 转换状态（BookingStatus -> LivingStatus）
                                if "status" in wecom_data:
                                    status_map = {
                                        BookingStatus.RESERVING: LivingStatus.RESERVED,
                                        BookingStatus.LIVING: LivingStatus.LIVING,
                                        BookingStatus.ENDED: LivingStatus.ENDED,
                                        BookingStatus.EXPIRED: LivingStatus.EXPIRED,
                                        BookingStatus.CANCELLED: LivingStatus.CANCELLED
                                    }
                                    wecom_data["status"] = status_map.get(wecom_data["status"], LivingStatus.RESERVED)
                                
                                final_data.update(wecom_data)
                            
                            # 检查是否存在
                            exists = session.query(Living).filter_by(livingid=livingid).first()
                            
                            if exists:
                                # 处理类型字段，确保它是 LivingType 枚举类型
                                type_value = final_data.get("type")
                                if type_value is not None:
                                    if isinstance(type_value, int) or (isinstance(type_value, str) and type_value.isdigit()):
                                        type_value = int(type_value)
                                        # 将数字映射到 LivingType 枚举
                                        type_map = {
                                            0: LivingType.GENERAL,
                                            1: LivingType.SMALL,
                                            2: LivingType.LARGE,
                                            3: LivingType.TRAINING,
                                            4: LivingType.EVENT
                                        }
                                        exists.type = type_map.get(type_value, LivingType.GENERAL)
                                
                                # 处理状态字段，确保它是 LivingStatus 枚举类型
                                status_value = final_data.get("status")
                                if status_value is not None:
                                    if isinstance(status_value, int) or (isinstance(status_value, str) and status_value.isdigit()):
                                        status_value = int(status_value)
                                        # 将数字映射到 LivingStatus 枚举
                                        status_map = {
                                            0: LivingStatus.RESERVED,
                                            1: LivingStatus.LIVING,
                                            2: LivingStatus.ENDED,
                                            3: LivingStatus.EXPIRED,
                                            4: LivingStatus.CANCELLED
                                        }
                                        exists.status = status_map.get(status_value, LivingStatus.RESERVED)
                                
                                # 更新现有记录的其他字段
                                exists.theme = final_data.get("theme", exists.theme)
                                exists.living_start = final_data.get("living_start", exists.living_start)
                                exists.living_duration = final_data.get("living_duration", exists.living_duration)
                                exists.anchor_userid = final_data.get("anchor_userid", exists.anchor_userid)
                                exists.description = final_data.get("description", exists.description)
                                # type 和 status 已在上面处理，这里不再赋值
                                
                                # 更新企业信息字段
                                if final_data.get("corpname"):
                                    exists.corpname = final_data.get("corpname")
                                elif not exists.corpname and livingid in livingid_user_map:
                                    # 如果corpname为空且知道该直播从哪个用户获取
                                    user_id = livingid_user_map[livingid]
                                    if user_id in user_objects and user_objects[user_id].corpname:
                                        exists.corpname = user_objects[user_id].corpname
                                
                                if final_data.get("agentid"):
                                    exists.agentid = final_data.get("agentid")
                                elif not exists.agentid and livingid in livingid_user_map:
                                    # 如果agentid为空且知道该直播从哪个用户获取
                                    user_id = livingid_user_map[livingid]
                                    if user_id in user_objects and user_objects[user_id].agentid:
                                        exists.agentid = user_objects[user_id].agentid
                                
                                exists.viewer_num = final_data.get("viewer_num", exists.viewer_num)
                                exists.comment_num = final_data.get("comment_num", exists.comment_num)
                                exists.mic_num = final_data.get("mic_num", exists.mic_num)
                                exists.online_count = final_data.get("online_count", exists.online_count)
                                exists.subscribe_count = final_data.get("subscribe_count", exists.subscribe_count)
                                # 设置远程同步状态为已同步
                                exists.is_remote_synced = 1
                                # 设置关联的预约ID
                                exists.live_booking_id = booking_id
                                updated_count += 1
                                
                                # 更新相关的LiveViewer记录的live_booking_id
                                if booking_id:
                                    session.query(LiveViewer).filter_by(living_id=exists.id).update(
                                        {"live_booking_id": booking_id}
                                    )
                            else:
                                # 创建新记录
                                # 处理类型字段，确保它是 LivingType 枚举类型
                                type_value = final_data.get("type")
                                if type_value is not None:
                                    if isinstance(type_value, int) or (isinstance(type_value, str) and type_value.isdigit()):
                                        type_value = int(type_value)
                                        # 将数字映射到 LivingType 枚举
                                        type_map = {
                                            0: LivingType.GENERAL,
                                            1: LivingType.SMALL,
                                            2: LivingType.LARGE,
                                            3: LivingType.TRAINING,
                                            4: LivingType.EVENT
                                        }
                                        type_value = type_map.get(type_value, LivingType.GENERAL)
                                else:
                                    type_value = LivingType.GENERAL
                                    
                                # 处理状态字段，确保它是 LivingStatus 枚举类型
                                status_value = final_data.get("status", LivingStatus.RESERVED)
                                if isinstance(status_value, int) or (isinstance(status_value, str) and status_value.isdigit()):
                                    status_value = int(status_value)
                                    # 将数字映射到 LivingStatus 枚举
                                    status_map = {
                                        0: LivingStatus.RESERVED,
                                        1: LivingStatus.LIVING,
                                        2: LivingStatus.ENDED,
                                        3: LivingStatus.EXPIRED,
                                        4: LivingStatus.CANCELLED
                                    }
                                    status_value = status_map.get(status_value, LivingStatus.RESERVED)
                                
                                # 获取企业信息，优先使用API数据，其次使用关联用户的企业信息
                                corpname = final_data.get("corpname", "")
                                agentid = final_data.get("agentid", "")
                                
                                # 如果企业信息为空，尝试从用户获取
                                if (not corpname or not agentid) and livingid in livingid_user_map:
                                    user_id = livingid_user_map[livingid]
                                    if user_id in user_objects:
                                        if not corpname and user_objects[user_id].corpname:
                                            corpname = user_objects[user_id].corpname
                                        if not agentid and user_objects[user_id].agentid:
                                            agentid = user_objects[user_id].agentid
                                
                                new_living = Living(
                                    livingid=livingid,
                                    theme=final_data.get("theme", ""),
                                    living_start=final_data.get("living_start", datetime.now()),
                                    living_duration=final_data.get("living_duration", 0),
                                    anchor_userid=final_data.get("anchor_userid", ""),
                                    description=final_data.get("description", ""),
                                    type=type_value,
                                    status=status_value,
                                    corpname=corpname,
                                    agentid=agentid,
                                    viewer_num=final_data.get("viewer_num", 0),
                                    comment_num=final_data.get("comment_num", 0),
                                    mic_num=final_data.get("mic_num", 0),
                                    online_count=final_data.get("online_count", 0),
                                    subscribe_count=final_data.get("subscribe_count", 0),
                                    is_remote_synced=1,  # 设置远程同步状态为已同步
                                    live_booking_id=booking_id  # 设置关联的预约ID
                                )
                                session.add(new_living)
                                created_count += 1
                    except Exception as e:
                        logger.error(f"同步直播数据失败 (ID: {livingid}): {str(e)}")
                
                # 提交所有更改
                session.commit()
            
            # 显示结果
            ErrorHandler.handle_info(
                f"同步直播数据完成\n更新记录：{updated_count}条\n新增记录：{created_count}条",
                self,
                "同步完成"
            )
            
            # 刷新数据
            self.load_data()
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, "同步直播数据失败")
            
    @PerformanceManager.measure_operation("cancel_live")
    def cancel_live(self, live: Living):
        """取消直播
        
        Args:
            live: 直播信息
        """
        try:
            # 确认取消
            if not ErrorHandler.handle_question(
                "确定要取消该直播吗？",
                self,
                "确认取消"
            ):
                return
                
            # 取消直播
            response = self.wecom_api.cancel_living(live.livingid)
            
            if response.get("errcode") == 0:
                # 更新直播状态
                with self.db_manager.get_session() as session:
                    # 从数据库中获取最新的实体并更新
                    stmt = select(Living).where(Living.livingid == live.livingid)
                    result = session.execute(stmt)
                    db_live = result.scalar_one_or_none()
                    if db_live:
                        db_live.status = LivingStatus.CANCELLED
                
                ErrorHandler.handle_info("取消直播成功", self, "成功")
                self.load_data()
                
            else:
                ErrorHandler.handle_warning(
                    f"取消直播失败：{response.get('errmsg')}",
                    self,
                    "失败"
                )
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "取消直播失败")
            
    @PerformanceManager.measure_operation("import_sign")
    def import_sign(self, live: Living):
        """
        导入签到数据
        :param live: 直播信息
        """
        try:
            # 先在一个会话中执行检查
            with self.db_manager.get_session() as session:
                # 重新获取直播记录
                live_record = session.query(Living).filter_by(livingid=live.livingid).first()
                if not live_record:
                    ErrorHandler.handle_warning("找不到直播记录", self, "错误")
                    return
                
                # 添加调试日志
                logger.info(f"找到直播记录：id={live_record.id}, livingid={live_record.livingid}")
                
                # 检查是否已拉取观看信息
                if not live_record.is_viewer_fetched:
                    ErrorHandler.handle_warning("请先拉取观看信息后再导入签到数据", self, "错误")
                    return
            
            # 显示文件选择对话框
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "选择签到文件",
                "",
                "Excel Files (*.xlsx *.xls);;CSV Files (*.csv);;All Files (*)"
            )
            
            if not file_path:
                return  # 用户取消了选择，直接返回
            
            # 使用进度对话框模式创建IO对话框
            io_dialog = IODialog(parent=self, title="导入签到记录", is_progress_dialog=True)
            io_dialog.add_info(f"正在导入签到数据，请稍候...")
            io_dialog.add_info(f"选择的文件: {file_path}")
            io_dialog.show()
            
            # 读取文件
            if file_path.endswith('.xlsx') or file_path.endswith('.xls'):
                try:
                    df = pd.read_excel(file_path)
                except Exception as e:
                    io_dialog.add_error(f"读取Excel文件失败: {str(e)}")
                    io_dialog.finish()  # 设置完成状态，但不关闭对话框
                    io_dialog.exec()  # 使用exec()而不是show()，等待用户关闭
                    return
            elif file_path.endswith('.csv'):
                try:
                    df = pd.read_csv(file_path)
                except Exception as e:
                    io_dialog.add_error(f"读取CSV文件失败: {str(e)}")
                    io_dialog.finish()
                    io_dialog.exec()
                    return
            else:
                io_dialog.add_error("不支持的文件格式")
                io_dialog.finish()
                io_dialog.exec()
                return
            
            # 创建SignImportManager实例
            from src.core.sign_import_manager import SignImportManager
            
            # 简化创建过程，移除不必要的企业信息获取
            import_manager = SignImportManager(self.db_manager)
            
            try:
                # 在一个会话中执行整个导入过程，以避免会话分离问题
                with self.db_manager.get_session() as session:
                    # 重新获取直播记录
                    live_record = session.query(Living).filter_by(livingid=live.livingid).first()
                    if not live_record:
                        io_dialog.add_error("找不到直播记录")
                        io_dialog.finish()
                        io_dialog.exec()
                        return
                    
                    # 添加进度信息
                    io_dialog.add_info(f"开始导入直播 \"{live_record.theme}\" 的签到数据...")
                    io_dialog.add_info(f"直播ID: {live_record.livingid}")
                    io_dialog.add_info(f"直播时间: {live_record.living_start}")
                    
                    # 导入签到数据，传递 live_record.id，确保它仍然与会话关联
                    import_results = import_manager.import_sign_data(file_path, live_record.id)
                    
                    # 解析导入结果
                    success_count = import_results.get('success_count', 0)
                    error_count = import_results.get('error_count', 0)
                    skipped_count = import_results.get('skipped_count', 0)
                    success_details = import_results.get('success_details', [])
                    error_details = import_results.get('error_details', [])
                    skipped_details = import_results.get('skipped_details', [])
                    
                    # 设置签到导入标志
                    live_record.is_sign_imported = 1
                    session.commit()
            except Exception as e:
                io_dialog.add_error(f"导入过程中发生错误: {str(e)}")
                import traceback
                io_dialog.add_error(f"错误详情: {traceback.format_exc()}")
                io_dialog.finish()
                io_dialog.exec()
                return
            
            # 显示导入结果
            io_dialog.add_info("\n===== 导入过程完成 =====")
            
            # 显示成功记录
            if success_count > 0:
                io_dialog.add_success(f"✓ 成功导入 {success_count} 条签到记录")
                # 显示成功详情（限制显示数量以避免过多）
                max_success_details = 10
                if success_details:
                    io_dialog.add_info(f"成功详情（显示前{min(len(success_details), max_success_details)}条，共{len(success_details)}条）:")
                    for detail in success_details[:max_success_details]:
                        io_dialog.add_success(f"  ✓ {detail}")
                    if len(success_details) > max_success_details:
                        io_dialog.add_info(f"  ... 还有 {len(success_details) - max_success_details} 条成功记录未显示")
            else:
                io_dialog.add_warning("没有成功导入的记录")
            
            # 显示跳过记录
            if skipped_count > 0:
                io_dialog.add_warning(f"⚠ 跳过 {skipped_count} 条重复记录")
                # 显示跳过详情
                if skipped_details:
                    max_skipped_details = 10
                    io_dialog.add_warning(f"跳过详情（显示前{min(len(skipped_details), max_skipped_details)}条，共{len(skipped_details)}条）:")
                    for detail in skipped_details[:max_skipped_details]:
                        io_dialog.add_warning(f"  ⚠ {detail}")
                    if len(skipped_details) > max_skipped_details:
                        io_dialog.add_warning(f"  ... 还有 {len(skipped_details) - max_skipped_details} 条跳过记录未显示")
            
            # 显示错误记录
            if error_count > 0:
                io_dialog.add_error(f"✗ 导入失败 {error_count} 条记录")
                # 显示错误详情
                if error_details:
                    max_error_details = 10
                    io_dialog.add_error(f"错误详情（显示前{min(len(error_details), max_error_details)}条，共{len(error_details)}条）:")
                    for detail in error_details[:max_error_details]:
                        io_dialog.add_error(f"  ✗ {detail}")
                    if len(error_details) > max_error_details:
                        io_dialog.add_error(f"  ... 还有 {len(error_details) - max_error_details} 条错误记录未显示")
            
            # 添加总结信息
            io_dialog.add_info("\n===== 导入结果总结 =====")
            io_dialog.add_info(f"总处理记录: {success_count + error_count + skipped_count}")
            io_dialog.add_success(f"成功导入: {success_count}")
            io_dialog.add_warning(f"跳过记录: {skipped_count}")
            io_dialog.add_error(f"导入失败: {error_count}")
            
            # 完成导入过程，但不自动关闭对话框
            io_dialog.finish()
            
            # 使用exec()方法显示对话框，并等待用户手动关闭
            io_dialog.exec()
            
            # 刷新数据
            self.load_data()
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, "导入签到失败")
            
    @PerformanceManager.measure_operation("fetch_watch_stat")
    def fetch_watch_stat(self, live: Living):
        """拉取直播观看信息"""
        logger.info("开始拉取直播观看信息")
        
        try:
            # 显示确认对话框
            confirm_box = QMessageBox(self)
            confirm_box.setIcon(QMessageBox.Icon.Question)
            confirm_box.setWindowTitle("确认拉取")
            confirm_box.setText(f"确定要拉取直播[{live.theme}]的观看信息吗？")
            confirm_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            confirm_box.setDefaultButton(QMessageBox.StandardButton.No)
            
            if confirm_box.exec() != QMessageBox.StandardButton.Yes:
                return
            
            # 显示进度对话框
            dialog = IODialog(parent=self, title="拉取观看信息", is_import=False, is_progress_dialog=True)
            dialog.show()
            dialog.add_info(f"正在拉取直播[{live.theme}]的观看信息...")
            
            # 创建LiveViewerManager实例
            from src.core.live_viewer_manager import LiveViewerManager
            viewer_manager = LiveViewerManager(
                self.db_manager,
                self.auth_manager
            )
            
            # 拉取观看信息 - 移除外部token参数，使用系统自动处理token
            success = viewer_manager.process_viewer_info(live.livingid)
            
            # 获取统计数据
            stats = viewer_manager.get_stats()
            
            # 检查是否成功拉取数据
            if not success:
                # 显示错误信息
                error_message = stats.get('last_error', '未知错误')
                dialog.add_error(f"观看数据拉取失败：{error_message}")
                dialog.finish()
                return
            
            # 检查是否有观众数据
            if stats.get('total_viewers', 0) == 0:
                dialog.add_warning("未获取到任何观看数据，请确认该直播有人观看")
                dialog.finish()
                
                # 尽管没有观众数据，仍然更新直播的拉取状态
                with self.db_manager.get_session() as session:
                    living_record = session.query(Living).filter_by(livingid=live.livingid).first()
                    if living_record:
                        living_record.is_viewer_fetched = 1
                        session.commit()
                        
                # 重新加载直播列表
                self.load_data()
                return
            
            # 显示成功信息
            dialog.add_success(f"观看数据拉取成功！")
            dialog.add_info(f"总观看人数: {stats.get('total_viewers', 0)}")
            dialog.add_info(f"内部成员: {stats.get('internal_viewers', 0)}")
            dialog.add_info(f"外部用户: {stats.get('external_viewers', 0)}")
                
            # 更新直播的拉取状态
            with self.db_manager.get_session() as session:
                living_record = session.query(Living).filter_by(livingid=live.livingid).first()
                if living_record:
                    living_record.is_viewer_fetched = 1
                    session.commit()
                    dialog.add_success("已更新拉取状态标记")
            
            # 重新加载直播列表
            self.load_data()
            
            # 完成操作
            dialog.finish()
            
            # 显示成功消息
            msg_box = AutoCloseMessageBox(
                "拉取成功", 
                f"已成功拉取直播[{live.theme}]的观看信息", 
                timeout=3000
            )
            msg_box.exec()
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "拉取观看信息失败")
            
    @PerformanceManager.measure_operation("export_data")
    def export_data(self):
        """导出数据"""
        try:
            # 创建默认文件名（包含当前日期时间）
            from datetime import datetime
            current_datetime = datetime.now().strftime("%Y%m%d_%H%M%S")
            default_filename = f"直播数据导出_{current_datetime}.xlsx"
            
            # 选择保存路径
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存Excel文件",
                default_filename,  # 使用默认文件名
                "Excel Files (*.xlsx)"
            )
            if not file_path:
                return
                
            # 获取所有直播记录
            with self.db_manager.get_session() as session:
                records = session.query(Living).all()
                
                if not records:
                    ErrorHandler.handle_warning("没有找到直播记录", self, "导出失败")
                    return
                    
                # 转换数据
                data = []
                
                for record in records:
                    # 创建记录数据字典
                    record_data = {}
                    
                    # 计算结束时间
                    end_time = ""
                    if record.living_start and record.living_duration:
                        try:
                            from datetime import timedelta
                            end_time_obj = record.living_start + timedelta(seconds=record.living_duration)
                            end_time = end_time_obj.strftime("%Y-%m-%d %H:%M:%S")
                        except Exception:
                            pass
                    
                    # 获取主播名称
                    anchor_name = record.anchor_userid
                    try:
                        user = session.query(User).filter(
                                (User.userid == record.anchor_userid) | (User.wecom_code == record.anchor_userid)
                        ).first()
                        if user:
                            anchor_name = f"{user.name}({record.anchor_userid})"
                    except Exception:
                        pass
                    
                    # 获取直播类型文本
                    from src.models.living import LivingType
                    type_text = {
                        LivingType.GENERAL: "通用直播",
                        LivingType.SMALL: "小班课",
                        LivingType.LARGE: "大班课",
                        LivingType.TRAINING: "企业培训",
                        LivingType.EVENT: "活动直播",
                    }.get(record.type, "未知")
                    
                    # 获取签到统计信息
                    from sqlalchemy import func
                    from src.models.live_viewer import LiveViewer
                    
                    with self.db_manager.get_session() as stats_session:
                        # 获取不同的签到人数
                        unique_signers = stats_session.query(func.count(LiveViewer.id)).filter(
                            LiveViewer.living_id == record.id,
                            LiveViewer.is_signed == True
                        ).scalar() or 0
                        
                        # 获取总签到次数
                        sign_count = stats_session.query(func.sum(LiveViewer.sign_count)).filter(
                            LiveViewer.living_id == record.id,
                            LiveViewer.is_signed == True
                        ).scalar() or 0
                        
                        # 获取首次签到时间
                        first_sign = stats_session.query(func.min(LiveViewer.sign_time)).filter(
                            LiveViewer.living_id == record.id,
                            LiveViewer.is_signed == True
                        ).scalar()
                        
                        sign_stats = {
                            "unique_signers": unique_signers,
                            "sign_count": sign_count,
                            "sign_time": first_sign
                        }
                    
                    # 使用中文字段名创建记录
                    data.append({
                        "直播ID": record.livingid,
                        "直播主题": record.theme,
                        "开始时间": record.living_start.strftime("%Y-%m-%d %H:%M:%S") if record.living_start else "",
                        "结束时间": end_time,
                        "直播时长(秒)": record.living_duration,
                        "主播": anchor_name,
                        "直播类型": type_text,
                        "状态": {
                            LivingStatus.RESERVED: "预约中",
                            LivingStatus.LIVING: "直播中",
                            LivingStatus.ENDED: "已结束",
                            LivingStatus.EXPIRED: "已过期",
                            LivingStatus.CANCELLED: "已取消"
                        }.get(record.status, "未知"),
                        "企业名称": record.corpname,
                        "描述": record.description or "",
                        "观看人数": record.viewer_num,
                        "评论数": record.comment_num,
                        "连麦人数": record.mic_num,
                        "在线人数": record.online_count,
                        "预约人数": record.subscribe_count,
                        "签到人数": sign_stats["unique_signers"],
                        "签到次数": sign_stats["sign_count"],
                        "首次签到时间": sign_stats["sign_time"].strftime("%Y-%m-%d %H:%M:%S") if sign_stats["sign_time"] else "",
                        "观看信息状态": "已拉取" if record.is_viewer_fetched == 1 else "未拉取",
                        "签到导入状态": "已导入" if record.is_sign_imported == 1 else "未导入",
                        "企微文档状态": "已上传" if record.is_doc_uploaded == 1 else "未上传",
                        "远程同步状态": "已同步" if record.is_remote_synced == 1 else "未同步",
                        "创建时间": record.created_at.strftime("%Y-%m-%d %H:%M:%S") if record.created_at else "",
                        "更新时间": record.updated_at.strftime("%Y-%m-%d %H:%M:%S") if record.updated_at else ""
                    })
                    
            # 创建DataFrame并导出到Excel
            df = pd.DataFrame(data)
            
            # 导出到Excel
            df.to_excel(file_path, index=False, sheet_name="直播数据")
            
            # 显示成功消息
            msg_box = AutoCloseMessageBox("导出成功", f"已成功导出 {len(data)} 条直播记录到文件:\n{file_path}", 3000, self)
            msg_box.exec()
            
        except Exception as e:
            ErrorHandler.handle_error(e, self, "导出数据失败")
            
    def refresh_data(self):
        """刷新数据"""
        self.load_data()
        
    def reset_search(self):
        """重置搜索条件"""
        self.live_title.clear()
        self.live_status.setCurrentIndex(0)
        
        # 重置状态字段
        self.viewer_fetched_status.setCurrentIndex(0) 
        self.sign_imported_status.setCurrentIndex(0)
        self.doc_uploaded_status.setCurrentIndex(0)
        self.remote_synced_status.setCurrentIndex(0)
        
        # 重置日期时间范围
        self.start_date_time.setDateTime(QDateTime.currentDateTime().addMonths(-1))
        self.end_date_time.setDateTime(QDateTime.currentDateTime())
        
        # 重置页码
        self.current_page = 1
        
        # 重新加载数据
        self.load_data()
    
    def show_combined_export_dialog(self):
        """显示综合导出数据对话框"""
        dialog = CombinedExportDialog(self.db_manager, self.auth_manager, self.user_id, self)
        dialog.exec()

    def _create_toolbar(self):
        """创建工具栏"""
        toolbar = QToolBar()
        toolbar.setObjectName("toolbar")
        return toolbar

class LiveDetailsDialog(QDialog):
    """直播详情对话框"""
    
    def __init__(self, live_info: Living, parent=None):
        super().__init__(parent)
        self.live_info = live_info  # 现在直接接收Living对象
        from src.core.database import DatabaseManager
        self.db_manager = DatabaseManager()  # 直接创建实例而不是调用get_db方法
        # 打印日志，检查状态值
        logger.debug(f"LiveDetailsDialog初始化 - 状态值: is_remote_synced={self.live_info.is_remote_synced}")
        self.init_ui()
        
    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle("直播详情")
        self.setMinimumWidth(800)  # 增加最小宽度，以适应左右布局
        self.setMinimumHeight(500)
        
        # 创建主布局（使用垂直布局）
        main_layout = QVBoxLayout(self)
        
        # 创建水平布局用于左右分布基本信息和统计信息
        top_layout = QHBoxLayout()
        main_layout.addLayout(top_layout)
        
        # 创建分组盒子 - 基本信息（左侧）
        basic_group = QGroupBox("基本信息")
        basic_layout = QFormLayout(basic_group)
        basic_layout.setSpacing(10)
        
        # 添加基本信息
        basic_layout.addRow("直播主题:", QLabel(self.live_info.theme))
        
        # 显示living_start和结束时间
        start_time_str = self.live_info.living_start.strftime("%Y-%m-%d %H:%M:%S") if self.live_info.living_start else ""
        basic_layout.addRow("开始时间:", QLabel(start_time_str))
        
        # 计算结束时间
        end_time = None
        if self.live_info.living_start and self.live_info.living_duration:
            end_time = self.live_info.living_start + timedelta(seconds=self.live_info.living_duration)
            end_time_str = end_time.strftime("%Y-%m-%d %H:%M:%S")
            basic_layout.addRow("结束时间:", QLabel(end_time_str))
        
        # 直播时长（转换为分钟和小时）
        if self.live_info.living_duration:
            hours = self.live_info.living_duration // 3600
            minutes = (self.live_info.living_duration % 3600) // 60
            seconds = self.live_info.living_duration % 60
            duration_str = ""
            if hours > 0:
                duration_str += f"{hours}小时"
            if minutes > 0:
                duration_str += f"{minutes}分钟"
            if seconds > 0 or duration_str == "":
                duration_str += f"{seconds}秒"
            basic_layout.addRow("直播时长:", QLabel(duration_str))
        
        # 主播信息
        try:
            from src.models.user import User
            with self.db_manager.get_session() as session:
                user = session.query(User).filter(
                    (User.wecom_code == self.live_info.anchor_userid) | 
                    (User.login_name == self.live_info.anchor_userid)
                ).first()
                if user:
                    anchor_name = f"{user.name}({self.live_info.anchor_userid})"
                else:
                    anchor_name = self.live_info.anchor_userid
                basic_layout.addRow("主播:", QLabel(anchor_name))
        except Exception as e:
            basic_layout.addRow("主播ID:", QLabel(self.live_info.anchor_userid))
            
        # 直播类型
        type_text = {
            LivingType.GENERAL: "通用直播",
            LivingType.SMALL: "小班课",
            LivingType.LARGE: "大班课",
            LivingType.TRAINING: "企业培训",
            LivingType.EVENT: "活动直播"
        }.get(self.live_info.type, "未知")
        basic_layout.addRow("直播类型:", QLabel(type_text))
        
        # 状态
        status_text = {
            LivingStatus.RESERVED: "预约中",
            LivingStatus.LIVING: "直播中",
            LivingStatus.ENDED: "已结束",
            LivingStatus.EXPIRED: "已过期",
            LivingStatus.CANCELLED: "已取消"
        }.get(self.live_info.status, "未知")
        basic_layout.addRow("状态:", QLabel(status_text))
        
        # 企业信息
        basic_layout.addRow("企业名称:", QLabel(self.live_info.corpname))
        
        # 描述
        if self.live_info.description:
            desc_label = QLabel(self.live_info.description)
            desc_label.setWordWrap(True)  # 允许文本自动换行
            basic_layout.addRow("直播描述:", desc_label)
            
        # 添加基本信息组到左侧
        top_layout.addWidget(basic_group)
        
        # 创建分组盒子 - 统计信息（右侧）
        stats_group = QGroupBox("统计信息")
        stats_layout = QFormLayout(stats_group)
        stats_layout.setSpacing(10)
        
        # 视频统计
        stats_layout.addRow("观看人数:", QLabel(str(self.live_info.viewer_num)))
        stats_layout.addRow("评论数:", QLabel(str(self.live_info.comment_num)))
        stats_layout.addRow("连麦人数:", QLabel(str(self.live_info.mic_num)))
        stats_layout.addRow("当前在线人数:", QLabel(str(self.live_info.online_count)))
        stats_layout.addRow("预约人数:", QLabel(str(self.live_info.subscribe_count)))
        
        # 签到统计
        if hasattr(self.live_info, 'sign_count'):
            stats_layout.addRow("签到人数:", QLabel(str(self.live_info.sign_count)))
        if hasattr(self.live_info, 'total_sign_count'):
            stats_layout.addRow("签到次数:", QLabel(str(self.live_info.total_sign_count)))
        if hasattr(self.live_info, 'sign_time') and self.live_info.sign_time:
            stats_layout.addRow("首次签到时间:", QLabel(str(self.live_info.sign_time)))
            
            # 计算首次签到相对于直播开始的时间
            if self.live_info.living_start:
                try:
                    from datetime import datetime
                    sign_time = datetime.strptime(self.live_info.sign_time, "%Y.%m.%d %H:%M")
                    time_diff = sign_time - self.live_info.living_start
                    minutes_diff = time_diff.total_seconds() // 60
                    if minutes_diff > 0:
                        stats_layout.addRow("首签距直播开始:", QLabel(f"{int(minutes_diff)}分钟后"))
                    elif minutes_diff < 0:
                        stats_layout.addRow("首签距直播开始:", QLabel(f"{int(abs(minutes_diff))}分钟前"))
                    else:
                        stats_layout.addRow("首签距直播开始:", QLabel("同时"))
                except Exception:
                    pass  # 忽略日期解析错误
        
        # 状态标记
        status_layout = QVBoxLayout()
        status_layout.setSpacing(5)
        
        # 是否拉取观看信息
        viewer_fetched_status = "已拉取" if self.live_info.is_viewer_fetched == 1 else "未拉取"
        viewer_label = QLabel(f"观看信息: {viewer_fetched_status}")
        viewer_label.setStyleSheet(f"color: {'green' if self.live_info.is_viewer_fetched == 1 else 'red'}")
        status_layout.addWidget(viewer_label)
        
        # 是否导入签到
        sign_imported_status = "已导入" if self.live_info.is_sign_imported == 1 else "未导入"
        sign_label = QLabel(f"签到导入: {sign_imported_status}")
        sign_label.setStyleSheet(f"color: {'green' if self.live_info.is_sign_imported == 1 else 'red'}")
        status_layout.addWidget(sign_label)
        
        # 是否上传企微文档
        doc_uploaded_status = "已上传" if self.live_info.is_doc_uploaded == 1 else "未上传"
        doc_label = QLabel(f"企微文档: {doc_uploaded_status}")
        doc_label.setStyleSheet(f"color: {'green' if self.live_info.is_doc_uploaded == 1 else 'red'}")
        status_layout.addWidget(doc_label)
        
        # 是否远程同步
        remote_synced_status = "已同步" if self.live_info.is_remote_synced == 1 else "未同步"
        remote_label = QLabel(f"远程同步: {remote_synced_status}")
        remote_label.setStyleSheet(f"color: {'green' if self.live_info.is_remote_synced == 1 else 'red'}")
        status_layout.addWidget(remote_label)
        
        # 添加状态布局
        stats_layout.addRow("数据状态:", status_layout)
        
        # 添加统计信息组到右侧
        top_layout.addWidget(stats_group)
        
        # 添加"查看详细观众信息"按钮
        details_btn_layout = QHBoxLayout()
        self.view_detail_btn = QPushButton("查看详细观众信息")
        self.view_detail_btn.setObjectName("primaryButton")
        self.view_detail_btn.setMinimumHeight(40)  # 设置按钮高度更大一些
        self.view_detail_btn.clicked.connect(self.open_viewer_details)
        details_btn_layout.addWidget(self.view_detail_btn, alignment=Qt.AlignCenter)
        main_layout.addLayout(details_btn_layout)
        
        # 添加关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.setObjectName("secondaryButton")
        close_btn.clicked.connect(self.accept)
        main_layout.addWidget(close_btn, alignment=Qt.AlignCenter)
        
    def open_viewer_details(self):
        """打开观众详情页面"""
        # 先关闭当前对话框
        self.accept()
        
        # 打开详细的观众信息页面
        detail_page = LiveViewerDetailPage(self.live_info, parent=self.parent())
        detail_page.exec()

class LiveViewerDetailPage(QDialog):
    """直播观众详情页面"""
    
    def __init__(self, live_info: Living, parent=None):
        super().__init__(parent)
        self.live_info = live_info
        self.db_manager = DatabaseManager()
        
        # 添加表头筛选条件
        self.header_filters = {}
        
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        """初始化UI"""
        self.setWindowTitle(f"观众详情 - {self.live_info.theme}")
        self.setMinimumWidth(1200)
        self.setMinimumHeight(800)
        
        main_layout = QVBoxLayout(self)
        
        # 1. 顶部信息区域 - 改为3列布局
        top_info_group = QGroupBox("直播信息")
        top_info_layout = QGridLayout(top_info_group)
        top_info_layout.setSpacing(10)
        
        # 获取主播名称
        anchor_name = "未知"
        with self.db_manager.get_session() as session:
            from src.models.user import User
            anchor = session.query(User).filter_by(userid=self.live_info.anchor_userid).first()
            if anchor:
                anchor_name = anchor.name
        
        # 计算直播结束时间
        end_time = ""
        if self.live_info.living_start and self.live_info.living_duration:
            from datetime import timedelta
            end_time = (self.live_info.living_start + timedelta(seconds=self.live_info.living_duration)).strftime('%Y-%m-%d %H:%M:%S')
        
        # 添加直播信息 - 使用网格布局，3列显示
        # 第一列
        top_info_layout.addWidget(QLabel("直播ID:"), 0, 0)
        top_info_layout.addWidget(QLabel(self.live_info.livingid), 0, 1)
        
        top_info_layout.addWidget(QLabel("直播主题:"), 1, 0)
        top_info_layout.addWidget(QLabel(self.live_info.theme), 1, 1)
        
        top_info_layout.addWidget(QLabel("直播时长:"), 2, 0)
        top_info_layout.addWidget(QLabel(f"{int(self.live_info.living_duration/60)} 分钟" if self.live_info.living_duration else "未知"), 2, 1)
        
        # 第二列
        top_info_layout.addWidget(QLabel("开始时间:"), 0, 2)
        top_info_layout.addWidget(QLabel(self.live_info.living_start.strftime('%Y-%m-%d %H:%M:%S') if self.live_info.living_start else "未知"), 0, 3)
        
        top_info_layout.addWidget(QLabel("结束时间:"), 1, 2)
        top_info_layout.addWidget(QLabel(end_time or "未知"), 1, 3)
        
        top_info_layout.addWidget(QLabel("主播:"), 2, 2)
        top_info_layout.addWidget(QLabel(f"{anchor_name} ({self.live_info.anchor_userid})"), 2, 3)
        
        # 第三列
        top_info_layout.addWidget(QLabel("直播类型:"), 0, 4)
        
        # 直播类型翻译
        living_type_map = {
            "GENERAL": "通用直播",
            "SMALL": "小班课",
            "LARGE": "大班课",
            "TRAINING": "企业培训",
            "EVENT": "活动直播",
        }
        live_type_text = living_type_map.get(str(self.live_info.type.name) if self.live_info.type else "GENERAL", "未知类型")
        top_info_layout.addWidget(QLabel(live_type_text), 0, 5)
        
        top_info_layout.addWidget(QLabel("观看人数:"), 1, 4)
        top_info_layout.addWidget(QLabel(str(self.live_info.viewer_num or 0)), 1, 5)
        
        top_info_layout.addWidget(QLabel("直播描述:"), 2, 4)
        top_info_layout.addWidget(QLabel(self.live_info.description or "无"), 2, 5)
        
        # 设置列的拉伸因子，使各列均匀分布
        for i in range(6):
            top_info_layout.setColumnStretch(i, 1)
        
        main_layout.addWidget(top_info_group)
        
        # 2. 搜索和筛选区域
        filter_group = QGroupBox("筛选条件")
        filter_layout = QVBoxLayout(filter_group)
        
        # 第一行筛选条件
        filter_row1 = QHBoxLayout()
        
        # 姓名筛选
        filter_row1.addWidget(QLabel("姓名:"))
        self.name_filter = QLineEdit()
        filter_row1.addWidget(self.name_filter)
        
        # 移除部门筛选
        
        # 签到状态筛选
        filter_row1.addWidget(QLabel("签到状态:"))
        self.sign_status_filter = QComboBox()
        self.sign_status_filter.addItems(["全部", "已签到", "未签到"])
        filter_row1.addWidget(self.sign_status_filter)
        
        # 观看时长筛选
        filter_row1.addWidget(QLabel("观看时长:"))
        self.watch_time_filter = QComboBox()
        self.watch_time_filter.addItems(["全部", "10分钟以下", "10-30分钟", "30-60分钟", "60分钟以上"])
        filter_row1.addWidget(self.watch_time_filter)
        
        filter_layout.addLayout(filter_row1)
        
        # 第二行筛选条件
        filter_row2 = QHBoxLayout()
        
        # 观看开始时间
        filter_row2.addWidget(QLabel("观看开始时间:"))
        self.watch_start_time_filter = CustomDateTimeWidget()
        current_time = QDateTime.currentDateTime()
        one_month_ago = current_time.addMonths(-1)
        self.watch_start_time_filter.setDateTime(one_month_ago)  # 默认为一个月前
        filter_row2.addWidget(self.watch_start_time_filter)
        
        filter_row2.addWidget(QLabel("至"))
        
        self.watch_end_time_filter = CustomDateTimeWidget()
        self.watch_end_time_filter.setDateTime(current_time)  # 默认为当前时间
        filter_row2.addWidget(self.watch_end_time_filter)
        
        # 将筛选按钮更名为搜索
        self.filter_btn = QPushButton("搜索")
        self.filter_btn.setObjectName("primaryButton")
        self.filter_btn.clicked.connect(self.apply_filter)
        filter_row2.addWidget(self.filter_btn)
        
        # 重置按钮
        self.reset_btn = QPushButton("重置")
        self.reset_btn.setObjectName("secondaryButton")
        self.reset_btn.clicked.connect(self.reset_filter)
        filter_row2.addWidget(self.reset_btn)
        
        # 导出按钮
        self.export_btn = QPushButton("导出数据")
        self.export_btn.setObjectName("primaryButton")
        self.export_btn.clicked.connect(self.export_data)
        filter_row2.addWidget(self.export_btn)
        
        filter_layout.addLayout(filter_row2)
        main_layout.addWidget(filter_group)
        
        # 3. 表格区域
        self.table = QTableWidget()
        # 初始化基础列，稍后再动态添加签到记录列
        self.table.setColumnCount(20)  # 更新列数以适应合并操作列
        self.table.setHorizontalHeaderLabels([
            "操作", "用户ID", "姓名", "用户来源", "用户类型", 
            "观看时长(分钟)", "是否评论", "是否连麦", "有无签到", "最后签到时间", "签到次数", 
            "邀请人ID", "邀请人姓名", "主播邀请", "位置信息", "符合奖励", "奖励金额", 
            "原始名称", "部门ID", "部门"
        ])
        
        # 设置表格样式
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setSelectionMode(QTableWidget.SingleSelection)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setAlternatingRowColors(True)
        
        # 使用自定义可筛选表头 - 必须先设置表头再启用排序
        custom_header = FilterHeader(self.table)
        self.table.setHorizontalHeader(custom_header)
        
        # 启用排序
        self.table.setSortingEnabled(True)
        
        # 设置列宽
        custom_header.setSectionResizeMode(QHeaderView.Interactive)
        
        # 基础列宽设置 - 增加列宽以适应筛选图标
        column_widths = [140, 100, 120, 100, 100, 120, 100, 100, 100, 170, 100, 
                        100, 120, 100, 160, 100, 100, 120, 140, 120, 140]
        for i, width in enumerate(column_widths):
            if i < len(column_widths):
                self.table.setColumnWidth(i, width)
        
        # 让一些列更宽一些
        self.table.setColumnWidth(0, 160)  # 操作列
        self.table.setColumnWidth(9, 190)  # 最后签到时间
        self.table.setColumnWidth(14, 160)  # 位置信息
        self.table.setColumnWidth(17, 120)  # 奖励状态
        
        main_layout.addWidget(self.table)
        
        # 4. 分页控件
        pagination_layout = QHBoxLayout()
        
        self.prev_btn = QPushButton("上一页")
        self.prev_btn.setObjectName("secondaryButton")
        self.prev_btn.clicked.connect(self.prev_page)
        pagination_layout.addWidget(self.prev_btn)
        
        self.page_label = QLabel("第 1 页 / 共 1 页")
        pagination_layout.addWidget(self.page_label, alignment=Qt.AlignCenter)
        
        self.next_btn = QPushButton("下一页")
        self.next_btn.setObjectName("secondaryButton")
        self.next_btn.clicked.connect(self.next_page)
        pagination_layout.addWidget(self.next_btn)
        
        main_layout.addLayout(pagination_layout)
        
        # 5. 按钮区域
        button_layout = QHBoxLayout()
        
        close_btn = QPushButton("关闭")
        close_btn.setObjectName("primaryButton")
        close_btn.clicked.connect(self.accept)
        button_layout.addWidget(close_btn, alignment=Qt.AlignCenter)
        
        main_layout.addLayout(button_layout)
        
        # 初始化分页变量
        self.current_page = 1
        self.page_size = 20  # 每页显示记录数
        self.total_pages = 1
        
    def load_data(self):
        """加载数据"""
        try:
            with self.db_manager.get_session() as session:
                # 获取筛选条件
                name_filter = self.name_filter.text().strip() if hasattr(self, 'name_filter') else ""
                # 移除部门筛选
                sign_status = self.sign_status_filter.currentText() if hasattr(self, 'sign_status_filter') else "全部"
                watch_time = self.watch_time_filter.currentText() if hasattr(self, 'watch_time_filter') else "全部"
                
                watch_start_time = None
                watch_end_time = None
                
                if hasattr(self, 'watch_start_time_filter') and hasattr(self, 'watch_end_time_filter'):
                    watch_start_time = self.watch_start_time_filter.dateTime().toPython() \
                        if not self.watch_start_time_filter.dateTime().isNull() else None
                    watch_end_time = self.watch_end_time_filter.dateTime().toPython() \
                        if not self.watch_end_time_filter.dateTime().isNull() else None
                
                # 基于 SQLAlchemy ORM 查询
                from src.models.live_viewer import LiveViewer
                from src.models.live_sign_record import LiveSignRecord
                from sqlalchemy import func, and_, or_, text, cast, String
                
                # 构建基础查询
                query = session.query(
                    LiveViewer.id,
                    LiveViewer.userid,
                    LiveViewer.name,
                    LiveViewer.user_source,
                    LiveViewer.user_type,
                    LiveViewer.department,
                    LiveViewer.department_id,
                    LiveViewer.watch_time,
                    LiveViewer.is_comment,
                    LiveViewer.is_mic,
                    LiveViewer.is_signed,
                    LiveViewer.sign_time,
                    LiveViewer.sign_count,
                    LiveViewer.invitor_userid,
                    LiveViewer.invitor_name,
                    LiveViewer.is_invited_by_anchor,
                    func.coalesce(cast(LiveViewer.location, String), '').label('location_info'),
                    LiveViewer.is_reward_eligible,
                    LiveViewer.reward_amount,
                    LiveViewer.reward_status  # 添加奖励状态字段
                ).filter(LiveViewer.living_id == self.live_info.id)
                
                # 应用姓名过滤
                if name_filter:
                    query = query.filter(LiveViewer.name.like(f"%{name_filter}%"))
                
                # 应用表头筛选条件 - 添加到后端查询中
                if hasattr(self, 'table') and hasattr(self.table, 'horizontalHeader') and isinstance(self.table.horizontalHeader(), FilterHeader):
                    header = self.table.horizontalHeader()
                    
                    # 用户ID筛选
                    if 1 in header.filters and header.filters[1]:
                        query = query.filter(LiveViewer.userid.in_(header.filters[1]))
                    elif 1 in header.custom_filters and header.custom_filters[1]:
                        query = query.filter(LiveViewer.userid.like(f"%{header.custom_filters[1]}%"))
                    
                    # 姓名筛选
                    if 2 in header.filters and header.filters[2]:
                        query = query.filter(LiveViewer.name.in_(header.filters[2]))
                    elif 2 in header.custom_filters and header.custom_filters[2]:
                        query = query.filter(LiveViewer.name.like(f"%{header.custom_filters[2]}%"))
                    
                    # 用户类型筛选
                    if 4 in header.filters and header.filters[4]:
                        user_types = []
                        for type_text in header.filters[4]:
                            if type_text == "企业微信用户":
                                user_types.append(2)
                            elif type_text == "微信用户":
                                user_types.append(1)
                        if user_types:
                            query = query.filter(LiveViewer.user_type.in_(user_types))
                    
                    # 签到状态筛选
                    if 8 in header.filters and header.filters[8]:
                        is_signed_values = []
                        for status_text in header.filters[8]:
                            if status_text == "已签到":
                                is_signed_values.append(True)
                            elif status_text == "未签到":
                                is_signed_values.append(False)
                        if is_signed_values:
                            query = query.filter(LiveViewer.is_signed.in_(is_signed_values))
                
                # 应用签到状态过滤
                if sign_status == "已签到":
                    query = query.filter(LiveViewer.is_signed == True)
                elif sign_status == "未签到":
                    query = query.filter(LiveViewer.is_signed == False)
                
                # 应用观看时长过滤
                if watch_time == "10分钟以下":
                    query = query.filter(LiveViewer.watch_time < 10 * 60)  # 转换为秒
                elif watch_time == "10-30分钟":
                    query = query.filter(LiveViewer.watch_time >= 10 * 60, LiveViewer.watch_time < 30 * 60)
                elif watch_time == "30-60分钟":
                    query = query.filter(LiveViewer.watch_time >= 30 * 60, LiveViewer.watch_time < 60 * 60)
                elif watch_time == "60分钟以上":
                    query = query.filter(LiveViewer.watch_time >= 60 * 60)
                
                # 移除对不存在属性的过滤
                # if watch_start_time:
                #     query = query.filter(LiveViewer.watch_start_time >= watch_start_time)
                # if watch_end_time:
                #     query = query.filter(LiveViewer.watch_start_time <= watch_end_time)
                
                # 计算总记录数
                total_count = query.count()
                self.total_pages = max(1, (total_count + self.page_size - 1) // self.page_size)
                
                # 获取当前页的数据
                viewers = query.order_by(LiveViewer.watch_time.desc())\
                    .offset((self.current_page - 1) * self.page_size)\
                    .limit(self.page_size).all()
                
                # 查找最大签到次数，用于动态调整列数
                max_sign_count = session.query(func.max(LiveViewer.sign_count))\
                    .filter(LiveViewer.living_id == self.live_info.id).scalar() or 0
                
                # 准备查询观众的签到记录
                viewer_ids = [viewer.id for viewer in viewers]
                
                # 查询所有相关的签到记录
                sign_records = {}
                original_member_names = {}
                if viewer_ids:
                    records = session.query(
                        LiveSignRecord
                    ).filter(
                        LiveSignRecord.viewer_id.in_(viewer_ids),
                        LiveSignRecord.living_id == self.live_info.livingid
                    ).order_by(
                        LiveSignRecord.viewer_id,
                        LiveSignRecord.sign_sequence
                    ).all()
                    
                    # 按观众ID分组签到记录
                    for record in records:
                        if record.viewer_id not in sign_records:
                            sign_records[record.viewer_id] = []
                        sign_records[record.viewer_id].append(record)
                        
                        # 记录原始成员名称
                        if record.original_member_name and record.viewer_id not in original_member_names:
                            original_member_names[record.viewer_id] = record.original_member_name
                
                # 动态设置表格列数和表头
                base_columns = 21  # 基础列数改为21（增加奖励状态列）
                total_columns = base_columns + max_sign_count
                self.table.setColumnCount(total_columns)
                
                # 设置表头
                headers = [
                    "操作", "用户ID", "姓名", "用户来源", "用户类型", 
                    "观看时长(分钟)", "是否评论", "是否连麦", "有无签到", "最后签到时间", "签到次数", 
                    "邀请人ID", "邀请人姓名", "主播邀请", "位置信息", "符合奖励", "奖励金额", "奖励状态",
                    "原始名称", "部门ID", "部门"
                ]
                
                # 添加动态签到时间列
                for i in range(1, max_sign_count + 1):
                    headers.append(f"第{i}次签到时间")
                    self.table.setColumnWidth(base_columns + i - 1, 180)  # 设置签到时间列的宽度
                
                self.table.setHorizontalHeaderLabels(headers)
                
                # 更新表格
                self.table.setRowCount(len(viewers))
                
                for row, viewer in enumerate(viewers):
                    # 合并操作列 - 创建操作按钮布局
                    widget = QWidget()
                    layout = QHBoxLayout(widget)
                    layout.setContentsMargins(5, 2, 5, 2)
                    layout.setSpacing(5)
                    
                    # 添加"查看详情"按钮
                    detail_btn = QPushButton("详情")
                    detail_btn.setFixedWidth(50)
                    detail_btn.setProperty("row", row)
                    detail_btn.setProperty("viewer_id", viewer.id)
                    detail_btn.clicked.connect(lambda checked, r=row, v_id=viewer.id: self.show_viewer_detail(r, v_id))
                    
                    # 添加"查看所有签到"按钮
                    all_signs_btn = QPushButton("签到记录")
                    all_signs_btn.setFixedWidth(65)
                    all_signs_btn.setProperty("row", row)
                    all_signs_btn.setProperty("viewer_id", viewer.id)
                    all_signs_btn.clicked.connect(lambda checked, r=row, v_id=viewer.id: self.show_all_sign_records(r, v_id))
                    
                    layout.addWidget(detail_btn)
                    layout.addWidget(all_signs_btn)
                    self.table.setCellWidget(row, 0, widget)
                    
                    # 添加其他数据
                    self.table.setItem(row, 1, QTableWidgetItem(viewer.userid))
                    self.table.setItem(row, 2, QTableWidgetItem(viewer.name))
                    
                    # 用户来源 - 翻译为中文
                    user_source_map = {
                        "INTERNAL": "内部企业成员",
                        "EXTERNAL": "外部联系人",
                        "UNKNOWN": "未知"
                    }
                    user_source = user_source_map.get(str(viewer.user_source.name) if viewer.user_source else "UNKNOWN", "未知")
                    self.table.setItem(row, 3, QTableWidgetItem(user_source))
                    
                    # 用户类型
                    user_type = "企业微信用户" if viewer.user_type == 2 else "微信用户"
                    self.table.setItem(row, 4, QTableWidgetItem(user_type))
                    
                    # 计算观看时长（分钟）
                    watch_minutes = int((viewer.watch_time or 0) / 60)
                    self.table.setItem(row, 5, QTableWidgetItem(f"{watch_minutes}分钟"))
                    
                    # 评论和连麦状态
                    self.table.setItem(row, 6, QTableWidgetItem("是" if viewer.is_comment else "否"))
                    self.table.setItem(row, 7, QTableWidgetItem("是" if viewer.is_mic else "否"))
                    
                    # 签到状态
                    sign_status_item = QTableWidgetItem("已签到" if viewer.is_signed else "未签到")
                    sign_status_item.setForeground(Qt.green if viewer.is_signed else Qt.red)
                    self.table.setItem(row, 8, sign_status_item)
                    
                    # 最后签到时间
                    last_sign_time = viewer.sign_time.strftime("%Y-%m-%d %H:%M:%S") if viewer.sign_time else "-"
                    self.table.setItem(row, 9, QTableWidgetItem(last_sign_time))
                    
                    # 签到次数
                    self.table.setItem(row, 10, QTableWidgetItem(str(viewer.sign_count or 0)))
                    
                    # 邀请人ID和姓名
                    self.table.setItem(row, 11, QTableWidgetItem(viewer.invitor_userid or ""))
                    self.table.setItem(row, 12, QTableWidgetItem(viewer.invitor_name or ""))
                    
                    # 主播邀请
                    self.table.setItem(row, 13, QTableWidgetItem("是" if viewer.is_invited_by_anchor else "否"))
                    
                    # 位置信息
                    self.table.setItem(row, 14, QTableWidgetItem(viewer.location_info or ""))
                    
                    # 奖励信息
                    self.table.setItem(row, 15, QTableWidgetItem("是" if viewer.is_reward_eligible else "否"))
                    self.table.setItem(row, 16, QTableWidgetItem(str(viewer.reward_amount or 0)))
                    
                    # 奖励状态
                    self.table.setItem(row, 17, QTableWidgetItem(viewer.reward_status or "未设置"))
                    
                    # 原始成员名称
                    original_name = original_member_names.get(viewer.id, "")
                    self.table.setItem(row, 18, QTableWidgetItem(original_name))
                    
                    # 部门ID和部门移到最后
                    self.table.setItem(row, 19, QTableWidgetItem(viewer.department_id or ""))
                    self.table.setItem(row, 20, QTableWidgetItem(viewer.department or ""))
                    
                    # 填充签到记录
                    viewer_records = sign_records.get(viewer.id, [])
                    for record in viewer_records:
                        if 1 <= record.sign_sequence <= max_sign_count:
                            col_index = base_columns + record.sign_sequence - 1
                            sign_time_str = record.sign_time.strftime("%Y-%m-%d %H:%M:%S") if record.sign_time else ""
                            sign_item = QTableWidgetItem(sign_time_str)
                            sign_type = record.sign_type if record.sign_type else "未知"
                            sign_item.setToolTip(f"类型: {sign_type}\n备注: {record.sign_remark or '无'}")
                            self.table.setItem(row, col_index, sign_item)
                
                # 更新分页信息
                self.page_label.setText(f"第 {self.current_page} 页 / 共 {self.total_pages} 页")
                self.prev_btn.setEnabled(self.current_page > 1)
                self.next_btn.setEnabled(self.current_page < self.total_pages)
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "加载观众数据失败")
    
    def show_viewer_detail(self, row, viewer_id):
        """显示观众详细信息"""
        try:
            with self.db_manager.get_session() as session:
                from src.models.live_viewer import LiveViewer
                from src.models.live_sign_record import LiveSignRecord
                
                viewer = session.query(LiveViewer).filter(LiveViewer.id == viewer_id).first()
                if not viewer:
                    ErrorHandler.handle_error(Exception(f"找不到ID为{viewer_id}的观众记录"), self, "查看详情失败")
                    return
                
                records = session.query(LiveSignRecord).filter(
                    LiveSignRecord.viewer_id == viewer_id,
                    LiveSignRecord.living_id == self.live_info.livingid
                ).order_by(LiveSignRecord.sign_sequence).all()
                
                # 构建详情信息
                detail_text = f"观众: {viewer.name} (ID: {viewer.userid})\n"
                detail_text += f"部门: {viewer.department or '未知'}\n"
                detail_text += f"观看时长: {int((viewer.watch_time or 0) / 60)}分钟\n"
                detail_text += f"签到状态: {'已签到' if viewer.is_signed else '未签到'}\n"
                detail_text += f"签到次数: {viewer.sign_count or 0}\n"
                detail_text += f"奖励资格: {'符合' if viewer.is_reward_eligible else '不符合'}\n"
                detail_text += f"奖励金额: {viewer.reward_amount or 0}\n"
                detail_text += f"奖励状态: {viewer.reward_status or '未设置'}\n\n"
                
                if records:
                    detail_text += "签到记录:\n"
                    for i, record in enumerate(records):
                        sign_time = record.sign_time.strftime("%Y-%m-%d %H:%M:%S") if record.sign_time else "未知"
                        detail_text += f"第{record.sign_sequence}次签到 - {sign_time}\n"
                        detail_text += f"签到类型: {record.sign_type or '未知'}\n"
                        if record.sign_remark:
                            detail_text += f"备注: {record.sign_remark}\n"
                        if i < len(records) - 1:
                            detail_text += "\n"
                else:
                    detail_text += "暂无签到记录"
                
                # 显示详情对话框
                detail_dialog = QDialog(self)
                detail_dialog.setWindowTitle(f"观众 {viewer.name} 详情")
                detail_dialog.setMinimumWidth(500)
                detail_dialog.setMinimumHeight(400)
                
                layout = QVBoxLayout(detail_dialog)
                
                # 详情文本区域
                text_edit = QTextEdit()
                text_edit.setReadOnly(True)
                text_edit.setText(detail_text)
                layout.addWidget(text_edit)
                
                # 关闭按钮
                close_btn = QPushButton("关闭")
                close_btn.clicked.connect(detail_dialog.accept)
                layout.addWidget(close_btn, alignment=Qt.AlignCenter)
                
                detail_dialog.exec_()
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "查看观众详情失败")
    
    def apply_filter(self):
        """应用筛选条件"""
        self.current_page = 1  # 重置为第一页
        self.load_data()
    
    def reset_filter(self):
        """重置筛选条件"""
        self.name_filter.clear()
        self.sign_status_filter.setCurrentIndex(0)  # "全部"
        self.watch_time_filter.setCurrentIndex(0)  # "全部"
        
        # 重置日期时间
        current_time = QDateTime.currentDateTime()
        one_month_ago = current_time.addMonths(-1)
        self.watch_start_time_filter.setDateTime(one_month_ago)
        self.watch_end_time_filter.setDateTime(current_time)
        
        # 清除表头筛选条件
        header = self.table.horizontalHeader()
        if isinstance(header, FilterHeader):
            # 先保存所有列索引
            columns_to_clear = list(header.filters.keys()) + list(header.custom_filters.keys())
            # 清除每一列的筛选条件
            for col in columns_to_clear:
                header.clearFilters(col)
            # 清空筛选列记录
            header.filtered_columns.clear()
            # 显示所有行
            for row in range(self.table.rowCount()):
                self.table.setRowHidden(row, False)
        
        self.current_page = 1  # 重置为第一页
        self.load_data()
    
    def prev_page(self):
        """上一页"""
        if self.current_page > 1:
            self.current_page -= 1
            self.load_data()
    
    def next_page(self):
        """下一页"""
        if self.current_page < self.total_pages:
            self.current_page += 1
            self.load_data()
    
    def show_all_sign_records(self, row, viewer_id):
        """显示观众所有签到记录"""
        try:
            with self.db_manager.get_session() as session:
                from src.models.live_viewer import LiveViewer
                from src.models.live_sign_record import LiveSignRecord
                
                viewer = session.query(LiveViewer).filter(LiveViewer.id == viewer_id).first()
                if not viewer:
                    ErrorHandler.handle_error(Exception(f"找不到ID为{viewer_id}的观众记录"), self, "查看签到记录失败")
                    return
                
                # 查询所有签到记录并按序号排序
                records = session.query(LiveSignRecord).filter(
                    LiveSignRecord.viewer_id == viewer_id,
                    LiveSignRecord.living_id == self.live_info.livingid
                ).order_by(LiveSignRecord.sign_sequence).all()
                
                if not records:
                    ErrorHandler.handle_info(f"观众 {viewer.name} 暂无签到记录", self, "签到记录")
                    return
                
                # 创建对话框显示所有签到记录
                dialog = QDialog(self)
                dialog.setWindowTitle(f"观众 {viewer.name} 的签到记录")
                dialog.setMinimumWidth(800)
                dialog.setMinimumHeight(500)
                
                layout = QVBoxLayout(dialog)
                
                # 创建表格
                table = QTableWidget()
                table.setColumnCount(7)
                table.setHorizontalHeaderLabels([
                    "序号", "签到时间", "签到类型", "签到备注", "是否有效", "签到Sheet", "原始名称"
                ])
                
                # 设置表格样式
                table.setSelectionBehavior(QTableWidget.SelectRows)
                table.setEditTriggers(QTableWidget.NoEditTriggers)
                table.setAlternatingRowColors(True)
                
                # 设置列宽
                table.setColumnWidth(0, 60)
                table.setColumnWidth(1, 180)
                table.setColumnWidth(2, 100)
                table.setColumnWidth(3, 200)
                table.setColumnWidth(4, 80)
                table.setColumnWidth(5, 120)
                table.setColumnWidth(6, 150)
                
                # 填充数据
                table.setRowCount(len(records))
                for i, record in enumerate(records):
                    table.setItem(i, 0, QTableWidgetItem(str(record.sign_sequence)))
                    
                    sign_time = record.sign_time.strftime("%Y-%m-%d %H:%M:%S") if record.sign_time else "未知"
                    table.setItem(i, 1, QTableWidgetItem(sign_time))
                    
                    table.setItem(i, 2, QTableWidgetItem(record.sign_type or "未知"))
                    table.setItem(i, 3, QTableWidgetItem(record.sign_remark or ""))
                    
                    valid_item = QTableWidgetItem("有效" if record.is_valid else "无效")
                    valid_item.setForeground(Qt.green if record.is_valid else Qt.red)
                    table.setItem(i, 4, valid_item)
                    
                    table.setItem(i, 5, QTableWidgetItem(record.sheet_name or ""))
                    table.setItem(i, 6, QTableWidgetItem(record.original_member_name or ""))
                
                layout.addWidget(table)
                
                # 关闭按钮
                close_btn = QPushButton("关闭")
                close_btn.clicked.connect(dialog.accept)
                layout.addWidget(close_btn, alignment=Qt.AlignCenter)
                
                dialog.exec_()
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "查看签到记录失败")
    
    def export_data(self):
        """导出数据到Excel"""
        try:
            # 选择保存路径
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "保存观众数据",
                f"直播观众数据_{self.live_info.livingid}.xlsx",
                "Excel Files (*.xlsx);;All Files (*)"
            )
            
            if not file_path:
                return  # 用户取消
                
            # 显示进度对话框
            progress = QProgressDialog("正在导出数据...", "取消", 0, 100, self)
            progress.setWindowTitle("导出进度")
            progress.setWindowModality(Qt.WindowModal)
            progress.show()
            
            # 获取所有观众数据（不分页）
            with self.db_manager.get_session() as session:
                # 获取筛选条件
                name_filter = self.name_filter.text().strip()
                sign_status = self.sign_status_filter.currentText()
                watch_time = self.watch_time_filter.currentText()
                
                watch_start_time = self.watch_start_time_filter.dateTime().toPython() \
                    if not self.watch_start_time_filter.dateTime().isNull() else None
                watch_end_time = self.watch_end_time_filter.dateTime().toPython() \
                    if not self.watch_end_time_filter.dateTime().isNull() else None
                
                # 基于 SQLAlchemy ORM 查询
                from src.models.live_viewer import LiveViewer
                from src.models.live_sign_record import LiveSignRecord
                from sqlalchemy import func, and_, or_, text, cast, String
                
                # 构建基础查询
                query = session.query(
                    LiveViewer.id,
                    LiveViewer.userid,
                    LiveViewer.name,
                    LiveViewer.user_source,
                    LiveViewer.user_type,
                    LiveViewer.department,
                    LiveViewer.department_id,
                    LiveViewer.watch_time,
                    LiveViewer.is_comment,
                    LiveViewer.is_mic,
                    LiveViewer.is_signed,
                    LiveViewer.sign_time,
                    LiveViewer.sign_count,
                    LiveViewer.invitor_userid,
                    LiveViewer.invitor_name,
                    LiveViewer.is_invited_by_anchor,
                    func.coalesce(cast(LiveViewer.location, String), '').label('location_info'),
                    LiveViewer.is_reward_eligible,
                    LiveViewer.reward_amount,
                    LiveViewer.reward_status  # 添加奖励状态字段
                ).filter(LiveViewer.living_id == self.live_info.id)
                
                progress.setValue(10)
                
                # 应用姓名过滤
                if name_filter:
                    query = query.filter(LiveViewer.name.like(f"%{name_filter}%"))
                
                # 应用签到状态过滤
                if sign_status == "已签到":
                    query = query.filter(LiveViewer.is_signed == True)
                elif sign_status == "未签到":
                    query = query.filter(LiveViewer.is_signed == False)
                
                # 应用观看时长过滤
                if watch_time == "10分钟以下":
                    query = query.filter(LiveViewer.watch_time < 10 * 60)  # 转换为秒
                elif watch_time == "10-30分钟":
                    query = query.filter(LiveViewer.watch_time >= 10 * 60, LiveViewer.watch_time < 30 * 60)
                elif watch_time == "30-60分钟":
                    query = query.filter(LiveViewer.watch_time >= 30 * 60, LiveViewer.watch_time < 60 * 60)
                elif watch_time == "60分钟以上":
                    query = query.filter(LiveViewer.watch_time >= 60 * 60)
                
                # 移除对不存在属性的过滤
                # if watch_start_time:
                #     query = query.filter(LiveViewer.watch_start_time >= watch_start_time)
                # if watch_end_time:
                #     query = query.filter(LiveViewer.watch_start_time <= watch_end_time)
                
                # 计算总记录数
                total_count = query.count()
                self.total_pages = max(1, (total_count + self.page_size - 1) // self.page_size)
                
                # 获取当前页的数据
                viewers = query.order_by(LiveViewer.watch_time.desc())\
                    .offset((self.current_page - 1) * self.page_size)\
                    .limit(self.page_size).all()
                
                progress.setValue(30)
                
                # 查找最大签到次数
                max_sign_count = session.query(func.max(LiveViewer.sign_count))\
                    .filter(LiveViewer.living_id == self.live_info.id).scalar() or 0
                
                # 获取所有观众的ID
                viewer_ids = [viewer.id for viewer in viewers]
                
                # 查询所有相关的签到记录
                sign_records = {}
                original_member_names = {}
                if viewer_ids:
                    records = session.query(
                        LiveSignRecord
                    ).filter(
                        LiveSignRecord.viewer_id.in_(viewer_ids),
                        LiveSignRecord.living_id == self.live_info.livingid
                    ).order_by(
                        LiveSignRecord.viewer_id,
                        LiveSignRecord.sign_sequence
                    ).all()
                    
                    # 按观众ID分组签到记录
                    for record in records:
                        if record.viewer_id not in sign_records:
                            sign_records[record.viewer_id] = {}
                        sign_records[record.viewer_id][record.sign_sequence] = record
                        
                        # 记录原始成员名称
                        if record.original_member_name and record.viewer_id not in original_member_names:
                            original_member_names[record.viewer_id] = record.original_member_name
                
                # 转换为 pandas DataFrame
                import pandas as pd
                data = []
                
                for i, viewer in enumerate(viewers):
                    if progress.wasCanceled():
                        return
                    
                    # 计算进度
                    if i % 10 == 0:
                        progress_value = 50 + int((i / len(viewers)) * 40)
                        progress.setValue(progress_value)
                    
                    # 计算观看时长（分钟）
                    watch_minutes = int((viewer.watch_time or 0) / 60)
                    
                    # 创建基本记录
                    record = {
                        "用户ID": viewer.userid,
                        "姓名": viewer.name,
                        "用户来源": str(viewer.user_source.name) if viewer.user_source else "未知",
                        "用户类型": "企业微信用户" if viewer.user_type == 2 else "微信用户",
                        "部门": viewer.department or "",
                        "部门ID": viewer.department_id or "",
                        "观看时长(分钟)": watch_minutes,
                        "是否评论": "是" if viewer.is_comment else "否",
                        "是否连麦": "是" if viewer.is_mic else "否",
                        "是否签到": "是" if viewer.is_signed else "否",
                        "最后签到时间": viewer.sign_time.strftime("%Y-%m-%d %H:%M:%S") if viewer.sign_time else "",
                        "签到次数": viewer.sign_count or 0,
                        "邀请人ID": viewer.invitor_userid or "",
                        "邀请人姓名": viewer.invitor_name or "",
                        "主播邀请": "是" if viewer.is_invited_by_anchor else "否",
                        "位置信息": viewer.location_info or "",
                        "符合奖励": "是" if viewer.is_reward_eligible else "否",
                        "奖励金额": viewer.reward_amount or 0,
                        "奖励状态": viewer.reward_status or "未设置",
                        "原始名称": original_member_names.get(viewer.id, "")
                    }
                    
                    # 添加动态签到记录
                    viewer_sign_records = sign_records.get(viewer.id, {})
                    for seq in range(1, max_sign_count + 1):
                        if seq in viewer_sign_records:
                            sign_record = viewer_sign_records[seq]
                            sign_time = sign_record.sign_time.strftime("%Y-%m-%d %H:%M:%S") if sign_record.sign_time else ""
                            record[f"第{seq}次签到时间"] = sign_time
                            record[f"第{seq}次签到类型"] = sign_record.sign_type or ""
                            record[f"第{seq}次签到备注"] = sign_record.sign_remark or ""
                            record[f"第{seq}次签到是否有效"] = "是" if sign_record.is_valid else "否"
                        else:
                            record[f"第{seq}次签到时间"] = ""
                            record[f"第{seq}次签到类型"] = ""
                            record[f"第{seq}次签到备注"] = ""
                            record[f"第{seq}次签到是否有效"] = ""
                    
                    data.append(record)
                
                progress.setValue(50)
                
                progress.setValue(90)
                
                # 创建DataFrame并保存到Excel
                df = pd.DataFrame(data)
                df.to_excel(file_path, index=False)
                
                progress.setValue(100)
                
                ErrorHandler.handle_info(f"已成功导出 {len(viewers)} 条记录到文件：\n{file_path}", self, "导出成功")
                
        except Exception as e:
            ErrorHandler.handle_error(e, self, "导出数据失败") 

class FilterHeader(QHeaderView):
    """可筛选的表头"""
    
    def __init__(self, parent=None):
        super(FilterHeader, self).__init__(Qt.Horizontal, parent)
        self.setSectionsClickable(True)
        self.setHighlightSections(True)
        self.setSortIndicatorShown(True)
        
        # 绘制筛选图标的大小和边距 - 增加大小和边距
        self.filter_icon_size = 18  # 由14增加到18
        self.filter_icon_margin = 6  # 由4增加到6
        
        # 存储过滤器
        self.filters = {}
        # 存储自定义筛选文本
        self.custom_filters = {}
        
        # 存储有筛选条件的列
        self.filtered_columns = set()
        
        # 添加双击事件处理 - 当双击表头时也显示筛选菜单
        self.setMouseTracking(True)
        self.doubleClicked = False
    
    def mousePressEvent(self, event):
        """处理鼠标点击事件"""
        pos = event.position().toPoint()
        # 获取点击位置对应的列索引
        logical_index = self.logicalIndexAt(pos)
        
        # 判断是否点击了筛选图标
        if self.isFilterIconClicked(pos, logical_index):
            # 点击了筛选图标，显示筛选菜单
            self.showFilterMenu(logical_index)
        else:
            # 点击了表头的其他区域，调用父类处理（执行排序）
            super(FilterHeader, self).mousePressEvent(event)
    
    def mouseDoubleClickEvent(self, event):
        """处理鼠标双击事件"""
        self.doubleClicked = True
        pos = event.position().toPoint()
        logical_index = self.logicalIndexAt(pos)
        # 双击时直接显示筛选菜单
        self.showFilterMenu(logical_index)
        # 防止事件传递
        event.accept()
    
    def isFilterIconClicked(self, pos, logical_index):
        """判断是否点击了筛选图标 - 扩大点击区域"""
        if logical_index < 0:
            return False
            
        section_rect = self.rect()
        section_rect.setLeft(self.sectionViewportPosition(logical_index))
        section_rect.setWidth(self.sectionSize(logical_index))
        
        # 扩大点击区域: 筛选图标位于表头右侧1/3区域
        # 计算出更大的点击区域，使整个表头右侧约1/3的区域都能点击
        icon_width = self.filter_icon_size + 16
        icon_left = section_rect.right() - icon_width - 20  # 排序图标宽度增加
        
        icon_rect = QRect(
            icon_left,  # 筛选图标位于表头右侧
            section_rect.top(),  # 从表头顶部开始
            icon_width,  # 宽度包括整个图标和边距
            section_rect.height()  # 整个表头高度
        )
        
        # 添加调试信息
        print(f"点击位置: {pos.x()}, {pos.y()}, 图标区域: {icon_rect}, 点击结果: {icon_rect.contains(pos)}")
        
        return icon_rect.contains(pos)
    
    def paintSection(self, painter, rect, logical_index):
        """重写绘制表头部分的方法，添加筛选图标"""
        # 调用父类的绘制方法
        super(FilterHeader, self).paintSection(painter, rect, logical_index)
        
        # 绘制筛选图标 - 调整位置使其更明显
        icon_width = self.filter_icon_size + 4
        icon_left = rect.right() - icon_width - 22  # 增加与排序图标的距离
        
        icon_rect = QRect(
            icon_left,
            rect.top() + (rect.height() - self.filter_icon_size) // 2,
            self.filter_icon_size,
            self.filter_icon_size
        )
        
        # 保存画笔状态
        painter.save()
        
        # 设置画笔和笔刷 - 增强颜色对比度
        if logical_index in self.filtered_columns:
            # 有筛选条件时使用深色蓝色
            painter.setPen(QPen(QColor(0, 120, 215), 2))  # 增加线宽
            painter.setBrush(QBrush(QColor(0, 120, 215, 220)))  # 增加不透明度
        else:
            # 无筛选条件时使用更深的灰色
            painter.setPen(QPen(QColor(50, 50, 50), 2))  # 更深的灰色，增加线宽
            painter.setBrush(QBrush(QColor(50, 50, 50, 200)))  # 增加不透明度
        
        # 绘制漏斗形状的筛选图标
        painter.setRenderHint(QPainter.Antialiasing)
        
        # 漏斗顶部的矩形
        painter.drawRect(
            icon_rect.left() + 2,
            icon_rect.top() + 2,
            icon_rect.width() - 4,
            (icon_rect.height() - 4) * 0.4
        )
        
        # 漏斗下部的三角形
        points = [
            QPoint(icon_rect.left() + 2, icon_rect.top() + 2 + (icon_rect.height() - 4) * 0.4),
            QPoint(icon_rect.right() - 2, icon_rect.top() + 2 + (icon_rect.height() - 4) * 0.4),
            QPoint(icon_rect.left() + icon_rect.width() // 2, icon_rect.bottom() - 2)
        ]
        painter.drawPolygon(points)
        
        # 恢复画笔状态
        painter.restore()
    
    def sectionSizeHint(self, logicalIndex):
        """提供表头列大小的建议值，加上筛选图标所需的额外空间"""
        # 获取原始的大小建议
        size = super(FilterHeader, self).sectionSizeHint(logicalIndex)
        # 额外添加筛选图标的大小和边距
        size += self.filter_icon_size + 2 * self.filter_icon_margin
        return size
    
    def showFilterMenu(self, logicalIndex):
        """显示筛选菜单"""
        table = self.parent()
        
        # 创建菜单
        menu = QMenu(self)
        
        # 获取当前列所有唯一值
        values = set()
        for i in range(table.rowCount()):
            item = table.item(i, logicalIndex)
            if item and item.text():
                values.add(item.text())
        
        # 添加选项到菜单
        for value in sorted(values):
            action = menu.addAction(value)
            action.setCheckable(True)
            action.setChecked(value in self.filters.get(logicalIndex, []))
            action.triggered.connect(lambda checked, v=value, idx=logicalIndex: self.toggleFilter(idx, v))
        
        # 添加操作选项
        if values:
            menu.addSeparator()
            all_action = menu.addAction("全选")
            all_action.triggered.connect(lambda: self.clearFilters(logicalIndex))
            
            clear_action = menu.addAction("取消筛选")
            clear_action.triggered.connect(lambda: self.applyFilter(logicalIndex, []))
            
            # 添加自定义筛选选项
            menu.addSeparator()
            custom_action = menu.addAction("自定义筛选...")
            custom_action.triggered.connect(lambda: self.showCustomFilterDialog(logicalIndex))
        
        # 获取当前列的屏幕位置
        column_pos = self.sectionPosition(logicalIndex)
        header_height = self.height()
        viewport_pos = self.viewport().mapToGlobal(QPoint(column_pos + self.sectionSize(logicalIndex) // 2, header_height))
        
        # 显示菜单在当前列下方
        menu.exec_(viewport_pos)
    
    def showCustomFilterDialog(self, columnIndex):
        """显示自定义筛选对话框"""
        current_text = self.custom_filters.get(columnIndex, "")
        text, ok = QInputDialog.getText(
            self, 
            "自定义筛选", 
            "输入筛选条件（支持部分匹配）:", 
            QLineEdit.Normal, 
            current_text
        )
        
        if ok and text:
            self.custom_filters[columnIndex] = text
            self.filtered_columns.add(columnIndex)  # 标记为已筛选
            self.applyCustomFilter(columnIndex, text)
        elif ok and not text and columnIndex in self.custom_filters:
            # 如果用户清空了输入，则删除该列的自定义筛选
            del self.custom_filters[columnIndex]
            # 检查是否还有其他筛选条件
            if columnIndex not in self.filters or not self.filters[columnIndex]:
                self.filtered_columns.discard(columnIndex)  # 移除筛选标记
            self.applyCustomFilter(columnIndex, "")
    
    def applyCustomFilter(self, columnIndex, text):
        """应用自定义筛选"""
        table = self.parent()
        
        # 清除该列的常规筛选
        if columnIndex in self.filters:
            del self.filters[columnIndex]
        
        # 显示所有行
        for row in range(table.rowCount()):
            table.setRowHidden(row, False)
        
        # 如果没有筛选文本，则仅应用其他列的筛选
        if not text:
            self.applyFilters()
            return
        
        # 应用筛选文本
        for row in range(table.rowCount()):
            should_hide = True
            item = table.item(row, columnIndex)
            if item and text.lower() in item.text().lower():
                should_hide = False
            
            if should_hide:
                table.setRowHidden(row, True)
        
        # 应用其他列的筛选
        self.applyFilters(skip_column=columnIndex)
        
        # 重新加载数据
        self.reloadData()
        
        # 重绘表头，显示筛选状态
        self.viewport().update()
    
    def toggleFilter(self, columnIndex, value):
        """切换特定值的筛选状态"""
        # 清除该列的自定义筛选
        if columnIndex in self.custom_filters:
            del self.custom_filters[columnIndex]
        
        if columnIndex not in self.filters:
            self.filters[columnIndex] = []
        
        if value in self.filters[columnIndex]:
            self.filters[columnIndex].remove(value)
        else:
            self.filters[columnIndex].append(value)
        
        # 更新筛选状态标记
        if self.filters[columnIndex]:
            self.filtered_columns.add(columnIndex)
        else:
            self.filtered_columns.discard(columnIndex)
        
        # 应用筛选
        self.applyFilters()
        
        # 重新加载数据
        self.reloadData()
        
        # 重绘表头，显示筛选状态
        self.viewport().update()
    
    def applyFilter(self, columnIndex, values):
        """应用特定列的筛选"""
        table = self.parent()
        
        # 如果是空列表，表示取消筛选，先显示所有行
        if not values:
            for row in range(table.rowCount()):
                table.setRowHidden(row, False)
        
        # 清除该列的自定义筛选
        if columnIndex in self.custom_filters:
            del self.custom_filters[columnIndex]
            
        self.filters[columnIndex] = values.copy()
        
        # 更新筛选状态标记
        if self.filters[columnIndex]:
            self.filtered_columns.add(columnIndex)
        else:
            self.filtered_columns.discard(columnIndex)
            
        self.applyFilters()
        
        # 如果取消筛选，重新加载所有数据确保UI刷新
        if not values:
            parent = self.parent()
            while parent:
                if isinstance(parent, LiveViewerDetailPage):
                    QTimer.singleShot(100, parent.load_data)
                    break
                parent = parent.parent()
        else:
            # 有筛选条件则只更新现有显示
            self.reloadData()
        
        # 重绘表头，显示筛选状态
        self.viewport().update()
    
    def clearFilters(self, columnIndex):
        """清除特定列的所有筛选"""
        table = self.parent()
        
        # 先显示所有行，以确保过滤条件解除后显示所有数据
        for row in range(table.rowCount()):
            table.setRowHidden(row, False)
            
        if columnIndex in self.filters:
            del self.filters[columnIndex]
        if columnIndex in self.custom_filters:
            del self.custom_filters[columnIndex]
        
        # 移除筛选状态标记
        self.filtered_columns.discard(columnIndex)
        
        # 重新应用其他列的筛选条件
        self.applyFilters()
        
        # 重新加载数据 - 取消注释，让其重新加载所有数据
        parent = self.parent()
        while parent:
            if isinstance(parent, LiveViewerDetailPage):
                QTimer.singleShot(100, parent.load_data)
                break
            parent = parent.parent()
        
        # 重绘表头，显示筛选状态
        self.viewport().update()
    
    def applyFilters(self, skip_column=None):
        """应用所有筛选"""
        table = self.parent()
        
        # 对每个筛选列进行处理
        for column, allowed_values in self.filters.items():
            if skip_column is not None and column == skip_column:
                continue
                
            if not allowed_values:  # 空列表表示不筛选
                continue
                
            for row in range(table.rowCount()):
                if table.isRowHidden(row):
                    continue  # 跳过已经隐藏的行
                    
                item = table.item(row, column)
                if not item or item.text() not in allowed_values:
                    table.setRowHidden(row, True)
    
    def reloadData(self):
        """重新加载数据"""
        # 检查筛选状态
        has_filters = bool(self.filters) or bool(self.custom_filters)
        
        # 查找包含此表头的窗口
        parent = self.parent()
        while parent:
            # 如果父窗口是LiveViewerDetailPage，则重新加载数据
            if isinstance(parent, LiveViewerDetailPage):
                # 如果没有筛选条件，重新加载所有数据
                if not has_filters:
                    QTimer.singleShot(100, parent.load_data)
                break
            parent = parent.parent()

class CombinedExportDialog(QDialog):
    """综合导出数据对话框"""
    
    def __init__(self, db_manager: DatabaseManager, auth_manager=None, user_id=None, parent=None):
        super().__init__(parent)
        self.db_manager = db_manager
        self.auth_manager = auth_manager
        self.user_id = user_id
        
        # 获取当前用户是否为管理员
        self.is_admin = False
        self.corpname = None
        if user_id:
            with self.db_manager.get_session() as session:
                user = session.query(User).filter(User.userid == user_id).first()
                if user:
                    self.is_admin = user.is_admin or user.role == UserRole.ROOT_ADMIN.value
                    self.corpname = user.corpname
        
        # 跟踪已选择的直播ID
        self.selected_live_ids = set()
        # 标记是否正在加载数据，避免循环调用
        self.is_loading = False
        
        self.setWindowTitle("综合导出页面")
        self.setMinimumWidth(1200)  # 增加对话框宽度，从800增加到1200
        self.setMinimumHeight(600)
        
        self.init_ui()
        self.load_live_list()
        
    def init_ui(self):
        """初始化UI"""
        main_layout = QVBoxLayout(self)
        
        # 上半部分 - 设置区域
        setting_group = QGroupBox("奖励设置")
        setting_layout = QFormLayout(setting_group)
        
        # 计算批次标签
        self.batch_label = QLabel()
        
        # 奖励规则方式
        self.rule_type_combo = QComboBox()
        rule_types = [
            ("sign", "仅签到次数"),
            ("watch", "仅观看时长"),
            ("count", "仅观看场次"),
            ("sign-watch", "签到次数和观看时长"),
            ("sign-count", "签到次数和观看场次"),
            ("watch-count", "观看时长和观看场次"),
            ("all-or", "所有条件满足其一即可"),
            ("all-and", "所有条件必须全满足")
        ]
        
        for value, text in rule_types:
            self.rule_type_combo.addItem(text, value)
        
        # 设置默认选择为"所有条件必须全满足"
        default_index = self.rule_type_combo.findData("all-and")
        if default_index != -1:
            self.rule_type_combo.setCurrentIndex(default_index)
        
        self.rule_type_combo.currentIndexChanged.connect(self.update_batch_label)
        
        # 同一观众最少观看场次
        self.min_watch_count_spin = QSpinBox()
        self.min_watch_count_spin.setRange(1, 100)
        self.min_watch_count_spin.setValue(3)
        
        setting_layout.addRow("计算批次:", self.batch_label)
        setting_layout.addRow("奖励规则方式:", self.rule_type_combo)
        setting_layout.addRow("同一观众最少观看场次(合格):", self.min_watch_count_spin)
        
        main_layout.addWidget(setting_group)
        
        # 下半部分 - 表格区域
        table_group = QGroupBox("直播列表")
        table_layout = QVBoxLayout(table_group)
        
        self.table = QTableWidget()
        self.table.setEditTriggers(QTableWidget.EditTrigger.DoubleClicked)
        self.table.setAlternatingRowColors(True)
        
        # 设置表头
        self.table.setColumnCount(7)  # 恢复为7列，移除操作列
        self.table.setHorizontalHeaderLabels([
            "序号", "直播名称", "直播时间", "主播", "签到次数", "观看时长(秒)", "红包奖励(￥)"
        ])
        
        # 创建5行
        self.table.setRowCount(5)
        
        # 设置默认值
        for row in range(5):
            # 序号列
            self.table.setItem(row, 0, QTableWidgetItem(str(row + 1)))
            
            # 创建直播名称选择器（ComboBox）和清除按钮的容器
            name_widget = QWidget()
            name_layout = QHBoxLayout(name_widget)
            name_layout.setContentsMargins(0, 0, 0, 0)
            name_layout.setSpacing(2)
            
            live_combo = QComboBox()
            live_combo.setEditable(True)
            live_combo.lineEdit().setPlaceholderText("请选择或搜索直播")
            live_combo.currentIndexChanged.connect(lambda idx, r=row: self.on_live_selected(r, idx))
            
            clear_button = QPushButton("×")
            clear_button.setFixedSize(24, 24)
            clear_button.setToolTip("清除选择")
            clear_button.clicked.connect(lambda checked, r=row: self.clear_live_selection(r))
            clear_button.setEnabled(False)  # 初始时禁用
            
            name_layout.addWidget(live_combo, 1)  # 设置拉伸因子，占据主要空间
            name_layout.addWidget(clear_button, 0)  # 不拉伸
            
            self.table.setCellWidget(row, 1, name_widget)
            
            # 存储对组件的引用，方便后续访问
            name_widget.setProperty("combo", live_combo)
            name_widget.setProperty("clear_button", clear_button)
            
            # 直播时间列（只读）
            self.table.setItem(row, 2, QTableWidgetItem(""))
            
            # 主播列（只读）
            self.table.setItem(row, 3, QTableWidgetItem(""))
            
            # 签到次数列
            sign_count_spin = QSpinBox()
            sign_count_spin.setRange(0, 100)
            sign_count_spin.setValue(3)  # 默认3次
            self.table.setCellWidget(row, 4, sign_count_spin)
            
            # 观看时长列
            watch_time_spin = QSpinBox()
            watch_time_spin.setRange(0, 10000)
            watch_time_spin.setValue(3000)  # 默认3000秒
            self.table.setCellWidget(row, 5, watch_time_spin)
            
            # 红包奖励列
            reward_amount = 0.0
            if row == 0:
                reward_amount = 1.88
            elif row == 1:
                reward_amount = 1.00
            elif row == 2:
                reward_amount = 1.62
                
            reward_spin = QDoubleSpinBox()
            reward_spin.setRange(0.0, 10000.0)
            reward_spin.setDecimals(2)
            reward_spin.setValue(reward_amount)
            reward_spin.setPrefix("￥")
            self.table.setCellWidget(row, 6, reward_spin)
        
        # 调整列宽
        self.table.setColumnWidth(0, 60)   # 序号
        self.table.setColumnWidth(1, 300)  # 直播名称
        self.table.setColumnWidth(2, 180)  # 直播时间
        self.table.setColumnWidth(3, 150)  # 主播
        self.table.setColumnWidth(4, 120)  # 签到次数
        self.table.setColumnWidth(5, 150)  # 观看时长
        self.table.setColumnWidth(6, 150)  # 红包奖励
        
        table_layout.addWidget(self.table)
        main_layout.addWidget(table_group)
        
        # 底部按钮
        button_layout = QHBoxLayout()
        button_layout.addStretch()
        
        self.calculate_btn = QPushButton("计算奖励结果")
        self.export_btn = QPushButton("导出数据")
        self.close_btn = QPushButton("关闭")
        
        self.calculate_btn.clicked.connect(self.calculate_reward)
        self.export_btn.clicked.connect(self.composite_export_data)
        self.close_btn.clicked.connect(self.close)
        
        button_layout.addWidget(self.calculate_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(self.close_btn)
        
        main_layout.addLayout(button_layout)
        
        # 初始化计算批次标签
        self.update_batch_label()
        
    def update_batch_label(self):
        """更新计算批次标签"""
        from datetime import datetime
        now = datetime.now()
        time_str = now.strftime("%Y%m%d%H%M%S")
        rule_type = self.rule_type_combo.currentData()
        
        batch_id = f"{time_str}-{self.user_id or 'anonymous'}-{rule_type}"
        self.batch_label.setText(batch_id)
        
    def load_live_list(self):
        """加载直播列表"""
        # 设置加载标志，避免循环调用
        if self.is_loading:
            return
            
        self.is_loading = True
        
        try:
            with self.db_manager.get_session() as session:
                query = session.query(Living)
                
                # 如果不是管理员，只显示当前企业的直播
                if not self.is_admin and self.corpname:
                    query = query.filter(Living.corpname == self.corpname)
                    
                # 获取所有直播，后续在Python中进行自定义排序
                lives = query.all()
                
                # 导入日期处理相关模块
                from datetime import datetime, timedelta, time
                
                # 获取今天的日期（只有年月日，没有时分秒）
                today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
                
                # 自定义排序函数
                def custom_sort_key(live):
                    # 获取直播日期和时间
                    live_date = live.living_start.replace(hour=0, minute=0, second=0, microsecond=0)
                    live_time = live.living_start.time()
                    
                    # 计算日期差（今天为0，昨天为1，前天为2，以此类推）
                    days_diff = (today - live_date).days
                    
                    # 根据日期差分为三类：今天、过去、未来
                    if days_diff < 0:
                        # 未来的直播，给一个很大的优先级值（10000）放到最后
                        # 同时按照时间从早到晚排序未来的直播
                        return (10000 + abs(days_diff), live_time)
                    elif days_diff == 0:
                        # 今天的直播
                        return (0, live_time)
                    else:
                        # 过去的直播
                        return (days_diff, live_time)
                
                # 按自定义规则排序
                sorted_lives = sorted(lives, key=custom_sort_key)
                
                # 更新每一行的下拉框
                for row in range(self.table.rowCount()):
                    name_widget = self.table.cellWidget(row, 1)
                    if not isinstance(name_widget, QWidget):
                        continue
                        
                    combo = name_widget.property("combo")
                    if not isinstance(combo, QComboBox):
                        continue
                        
                    # 暂时断开信号，避免触发 on_live_selected
                    combo.blockSignals(True)
                    
                    combo.clear()
                    combo.addItem("", None)  # 空选项
                    
                    # 获取当前行已选择的直播ID（若有）
                    current_selection = None
                    if combo.property("selected_live_id"):
                        current_selection = combo.property("selected_live_id")
                    
                    for live in sorted_lives:
                        # 添加日期和时间信息到显示文本
                        date_str = live.living_start.strftime("%Y-%m-%d")
                        time_str = live.living_start.strftime("%H:%M")
                        display_text = f"{live.theme}({live.livingid})({date_str} {time_str})"
                        
                        # 如果直播已被选择且不是当前行选择的，则不添加
                        if live.id in self.selected_live_ids and live.id != current_selection:
                            continue
                            
                        combo.addItem(display_text, live.id)
                        
                        # 如果是当前行选择的直播，选中它
                        if live.id == current_selection:
                            combo.setCurrentIndex(combo.count() - 1)
                            
                    # 重新连接信号
                    combo.blockSignals(False)
                    
        except Exception as e:
            logger.error(f"加载直播列表时出错: {str(e)}")
            QMessageBox.critical(self, "错误", f"加载直播列表时出错: {str(e)}")
        
        # 重置加载标志
        self.is_loading = False
            
    def on_live_selected(self, row, index):
        """当选择直播时更新相关信息"""
        # 如果正在加载数据，忽略此次回调
        if self.is_loading:
            return
            
        name_widget = self.table.cellWidget(row, 1)
        if not isinstance(name_widget, QWidget):
            return
            
        combo = name_widget.property("combo")
        clear_button = name_widget.property("clear_button")
        
        if not isinstance(combo, QComboBox):
            return
            
        # 获取之前选择的直播ID（如果有）
        previous_live_id = combo.property("selected_live_id")
        if previous_live_id and previous_live_id in self.selected_live_ids:
            self.selected_live_ids.remove(previous_live_id)
            
        # 如果选择了空选项
        if index <= 0:
            combo.setProperty("selected_live_id", None)
            
            # 更新清除按钮状态
            if isinstance(clear_button, QPushButton):
                clear_button.setEnabled(False)
                
            # 清空直播时间和主播
            self.table.setItem(row, 2, QTableWidgetItem(""))
            self.table.setItem(row, 3, QTableWidgetItem(""))
            
            # 重新加载列表以更新可用选项
            self.load_live_list()
            return
            
        live_id = combo.itemData(index)
        if not live_id:
            return
            
        try:
            with self.db_manager.get_session() as session:
                live = session.query(Living).filter(Living.id == live_id).first()
                
                if live:
                    # 更新直播时间
                    time_item = QTableWidgetItem(live.living_start.strftime("%Y-%m-%d %H:%M:%S"))
                    self.table.setItem(row, 2, time_item)
                    
                    # 更新主播
                    anchor_item = QTableWidgetItem(live.anchor_userid)
                    self.table.setItem(row, 3, anchor_item)
                    
                    # 添加到已选择集合
                    self.selected_live_ids.add(live_id)
                    combo.setProperty("selected_live_id", live_id)
                    
                    # 启用清除按钮
                    if isinstance(clear_button, QPushButton):
                        clear_button.setEnabled(True)
                    
                    # 重新加载列表以更新可用选项
                    self.load_live_list()
            
        except Exception as e:
            logger.error(f"更新直播信息时出错: {str(e)}")

    def clear_live_selection(self, row):
        """清除直播选择"""
        name_widget = self.table.cellWidget(row, 1)
        if not isinstance(name_widget, QWidget):
            return
            
        combo = name_widget.property("combo")
        clear_button = name_widget.property("clear_button")
        
        if not isinstance(combo, QComboBox):
            return
            
        # 获取当前选择的直播ID
        live_id = combo.property("selected_live_id")
        if live_id and live_id in self.selected_live_ids:
            self.selected_live_ids.remove(live_id)
            
        # 重置组合框
        combo.setCurrentIndex(0)  # 选择空选项
        combo.setProperty("selected_live_id", None)
        
        # 清空直播时间和主播
        self.table.setItem(row, 2, QTableWidgetItem(""))
        self.table.setItem(row, 3, QTableWidgetItem(""))
        
        # 禁用清除按钮
        if isinstance(clear_button, QPushButton):
            clear_button.setEnabled(False)
            
        # 重新加载列表以更新可用选项
        self.load_live_list()

    def calculate_reward(self):
        """计算奖励结果"""
        from datetime import datetime
        from src.models.live_viewer import LiveViewer
        from src.models.live_sign_record import LiveSignRecord
        from src.models.living import Living
        
        logger.info("==================== 开始计算奖励 ====================")
        
        # 获取界面上的规则设置
        rule_type_str = self.rule_type_combo.currentData()
        batch_id = self.batch_label.text()
        
        logger.info(f"规则类型: {rule_type_str}, 批次ID: {batch_id}")
        
        # 转换规则类型
        try:
            rule_type = RewardRuleType.from_string(rule_type_str)
            logger.info(f"解析规则类型成功: {rule_type}, 对应值: {rule_type.value}")
        except Exception as e:
            rule_type = RewardRuleType.ALL_AND  # 默认值
            logger.warning(f"解析规则类型失败，使用默认值: {rule_type}, 错误: {str(e)}")
        
        # 收集所有选中的直播数据
        selected_lives = []
        batch_living_ids = []  # 存储所有直播ID，用于计算观看场次
        
        logger.info("开始收集选中的直播数据...")
        
        for row in range(self.table.rowCount()):
            name_widget = self.table.cellWidget(row, 1)
            if not isinstance(name_widget, QWidget):
                continue
                
            combo = name_widget.property("combo")
            if not isinstance(combo, QComboBox):
                continue
                
            # 获取选择的直播ID
            live_id = combo.property("selected_live_id")
            if not live_id:
                continue
                
            # 获取签到次数和观看时长规则
            sign_count_spin = self.table.cellWidget(row, 4)
            watch_time_spin = self.table.cellWidget(row, 5)
            reward_spin = self.table.cellWidget(row, 6)
            
            if not all([isinstance(sign_count_spin, QSpinBox), 
                       isinstance(watch_time_spin, QSpinBox),
                       isinstance(reward_spin, QDoubleSpinBox)]):
                continue
                
            rule_sign_count = sign_count_spin.value()
            rule_watch_time = watch_time_spin.value()
            reward_amount = reward_spin.value()
            
            logger.info(f"选中直播ID: {live_id}, 签到次数规则: {rule_sign_count}, 观看时长规则: {rule_watch_time}, 奖励金额: {reward_amount}")
            
            # 将数据添加到待处理列表
            selected_lives.append({
                'id': live_id,
                'rule_sign_count': rule_sign_count,
                'rule_watch_time': rule_watch_time,
                'reward_amount': reward_amount,
            })
            
            batch_living_ids.append(live_id)
            
        # 检查是否有选择的直播
        if not selected_lives:
            logger.warning("没有选择直播进行计算，退出计算过程")
            QMessageBox.warning(self, "警告", "请至少选择一个直播进行计算。")
            return
        
        logger.info(f"共选择了 {len(selected_lives)} 个直播进行计算")
            
        # 获取同一观众最少观看场次
        rule_watch_count = self.min_watch_count_spin.value()
        logger.info(f"最少观看场次规则: {rule_watch_count}")
        
        # 创建处理进度对话框
        progress_dialog = QProgressDialog("准备处理数据...", "取消", 0, 100, self)
        progress_dialog.setWindowModality(Qt.WindowModal)
        progress_dialog.setMinimumDuration(0)
        progress_dialog.setValue(0)
        progress_dialog.show()
        QApplication.processEvents()
        
        # 创建共享数据
        data_lock = Lock()
        total_viewers = 0
        processed_viewers = 0
        is_cancelled = False
        
        # 统一的时间戳，避免不同记录间的时间戳差异
        current_timestamp = datetime.now()
        logger.info(f"设置统一时间戳: {current_timestamp}")
        
        # 进度更新函数
        def update_progress(phase="处理中", value=None):
            nonlocal processed_viewers, total_viewers
            if value is not None:
                progress_dialog.setValue(value)
            elif total_viewers > 0 and not is_cancelled:
                progress = int((processed_viewers / total_viewers) * 100)
                progress_dialog.setValue(min(progress, 100))
            
            progress_dialog.setLabelText(f"{phase}...")
            QApplication.processEvents()
            return progress_dialog.wasCanceled()
        
        try:
            update_progress("正在获取直播数据", 5)
            
            # 查询并收集所有需要的数据 - 移出线程中避免数据库连接问题
            all_live_configs = {}
            all_viewers = {}
            all_sign_records = {}
            all_watch_counts = {}
            
            # 记录所有用户ID到直播ID的映射，用于计算场次
            user_id_to_live_id = defaultdict(list)
            
            # 使用主线程获取所有数据，避免子线程中出现数据库连接问题
            with self.db_manager.get_session() as session:
                # 1. 获取所有选中直播的观众和签到数据
                logger.info("开始查询所有选中直播的观众数据...")
                
                for live_config in selected_lives:
                    live_id = live_config['id']
                    all_live_configs[live_id] = live_config
                    
                    # 获取观众数据
                    viewers = session.query(LiveViewer).filter(LiveViewer.living_id == live_id).all()
                    all_viewers[live_id] = []
                    
                    for viewer in viewers:
                        # 记录用户ID到直播ID的映射
                        user_id_to_live_id[viewer.userid].append(live_id)
                        
                        # 转换为字典保存
                        viewer_dict = {
                            'id': viewer.id,
                            'userid': viewer.userid,
                            'name': viewer.name,
                            'watch_time': viewer.watch_time,
                            'is_signed': viewer.is_signed,
                            'sign_time': viewer.sign_time,
                            'sign_count': viewer.sign_count or 0
                        }
                        all_viewers[live_id].append(viewer_dict)
                    
                    logger.info(f"直播ID: {live_id} 的观众数量: {len(all_viewers[live_id])}")
                    total_viewers += len(all_viewers[live_id])
                
                logger.info(f"总观众数量: {total_viewers}")
                
                # 2. 获取所有观众的签到记录
                logger.info("开始查询所有观众的签到记录...")
                
                for live_id, viewers in all_viewers.items():
                    viewer_ids = [v['id'] for v in viewers]
                    
                    if not viewer_ids:
                        continue
                    
                    # 使用分批查询减少内存使用
                    batch_size = 500
                    all_sign_records[live_id] = {}
                    
                    for i in range(0, len(viewer_ids), batch_size):
                        batch_ids = viewer_ids[i:i+batch_size]
                        
                        # 使用分组查询获取签到次数
                        sign_counts = session.query(
                            LiveSignRecord.viewer_id, 
                            LiveSignRecord.living_id, 
                            func.count(LiveSignRecord.id).label('count')
                        ).filter(
                            LiveSignRecord.viewer_id.in_(batch_ids),
                            LiveSignRecord.living_id == live_id
                        ).group_by(
                            LiveSignRecord.viewer_id, 
                            LiveSignRecord.living_id
                        ).all()
                        
                        for viewer_id, living_id, count in sign_counts:
                            if viewer_id not in all_sign_records[live_id]:
                                all_sign_records[live_id][viewer_id] = {}
                            all_sign_records[live_id][viewer_id][str(living_id)] = count
                
                # 3. 计算每个观众的观看场次
                logger.info("开始计算每个观众的观看场次...")
                
                for userid, live_ids in user_id_to_live_id.items():
                    # 计算当前用户观看的不同直播数
                    unique_live_ids = set(live_ids)
                    watch_count = len(unique_live_ids)
                    all_watch_counts[userid] = watch_count
                    logger.debug(f"用户 {userid} 观看了 {watch_count} 场直播")
            
            update_progress("预处理数据完成", 10)
            
            # 准备线程处理函数
            def process_chunk(live_id, viewers_chunk, sign_records, watch_counts, live_config, rule_watch_count_value):
                nonlocal processed_viewers, is_cancelled, data_lock
                
                logger.debug(f"开始处理直播 {live_id} 的观众数据块，数量: {len(viewers_chunk)}")
                
                if is_cancelled:
                    return None
                
                # 从配置中获取规则设置
                rule_sign_count = live_config['rule_sign_count']
                # 修复这里的错误 - 移除将分钟转换为秒的代码
                rule_watch_time = live_config['rule_watch_time']  # UI传入的已经是秒，不需要再转换
                rule_amount = live_config['reward_amount']  # 页面设置的金额
                
                logger.debug(f"直播 {live_id} 的规则设置: 签到次数={rule_sign_count}, 观看时长={rule_watch_time}秒, 奖励金额={rule_amount}")
                
                # 准备结果数据结构 - 深拷贝避免共享数据问题
                reward_records = []
                viewer_updates = []
                
                # 处理每个观众
                for viewer in viewers_chunk:
                    if is_cancelled:
                        return None
                    
                    viewer_id = viewer['id']
                    userid = viewer['userid']
                    # 添加观众名称，方便日志识别
                    viewer_name = viewer['name']
                    
                    # 增加详细日志，清晰记录每个观众的原始数据
                    logger.info(f"处理观众: {viewer_name}(ID:{viewer_id}, UserID:{userid})")
                    logger.info(f"观众原始数据: 是否签到={viewer['is_signed']}, 观看时长={viewer['watch_time']}秒, 签到次数={viewer['sign_count']}")
                    
                    # 获取签到记录数 - 从预加载数据中获取
                    sign_count = 0
                    if viewer_id in sign_records and str(live_id) in sign_records[viewer_id]:
                        sign_count = sign_records[viewer_id][str(live_id)]
                    
                    # 修复：如果从签到记录中无法获取签到次数，则使用viewer本身的sign_count
                    if sign_count == 0 and viewer['is_signed'] and viewer['sign_count'] > 0:
                        sign_count = viewer['sign_count']
                        logger.info(f"从viewer记录获取的签到次数: {sign_count}")
                    
                    logger.info(f"观众 {viewer_name} 的签到次数: {sign_count}")
                    
                    # 获取观看次数 - 从预加载数据中获取
                    watch_count = 0
                    if userid in watch_counts:
                        watch_count = watch_counts[userid]
                    
                    logger.info(f"观众 {viewer_name} 的观看场次: {watch_count}")
                    
                    # 获取观看时长
                    watch_time = viewer['watch_time'] or 0
                    logger.info(f"观众 {viewer_name} 的观看时长: {watch_time}秒")
                    
                    # 添加更详细的规则比较日志
                    logger.info(f"规则要求: 签到次数≥{rule_sign_count}, 观看时长≥{rule_watch_time}秒, 观看场次≥{rule_watch_count_value}")
                    
                    # 判断是否符合奖励条件
                    is_eligible = False
                    
                    # 增加规则判断的详细日志
                    if rule_type == RewardRuleType.SIGN:
                        is_eligible = sign_count >= rule_sign_count
                        logger.info(f"观众 {viewer_name} 签到规则检查: {sign_count} >= {rule_sign_count} = {is_eligible}")
                    elif rule_type == RewardRuleType.WATCH:
                        is_eligible = watch_time >= rule_watch_time
                        logger.info(f"观众 {viewer_name} 观看时长规则检查: {watch_time} >= {rule_watch_time} = {is_eligible}")
                    elif rule_type == RewardRuleType.COUNT:
                        is_eligible = watch_count >= rule_watch_count_value
                        logger.info(f"观众 {viewer_name} 观看场次规则检查: {watch_count} >= {rule_watch_count_value} = {is_eligible}")
                    elif rule_type == RewardRuleType.SIGN_WATCH:
                        sign_check = sign_count >= rule_sign_count
                        watch_check = watch_time >= rule_watch_time
                        is_eligible = sign_check and watch_check
                        logger.info(f"观众 {viewer_name} 签到+观看时长规则检查: 签到({sign_count} >= {rule_sign_count} = {sign_check}) 且 观看时长({watch_time} >= {rule_watch_time} = {watch_check}) = {is_eligible}")
                    elif rule_type == RewardRuleType.SIGN_COUNT:
                        sign_check = sign_count >= rule_sign_count
                        count_check = watch_count >= rule_watch_count_value
                        is_eligible = sign_check and count_check
                        logger.info(f"观众 {viewer_name} 签到+观看场次规则检查: 签到({sign_count} >= {rule_sign_count} = {sign_check}) 且 观看场次({watch_count} >= {rule_watch_count_value} = {count_check}) = {is_eligible}")
                    elif rule_type == RewardRuleType.WATCH_COUNT:
                        watch_check = watch_time >= rule_watch_time
                        count_check = watch_count >= rule_watch_count_value
                        is_eligible = watch_check and count_check
                        logger.info(f"观众 {viewer_name} 观看时长+观看场次规则检查: 观看时长({watch_time} >= {rule_watch_time} = {watch_check}) 且 观看场次({watch_count} >= {rule_watch_count_value} = {count_check}) = {is_eligible}")
                    elif rule_type == RewardRuleType.ALL_OR:
                        sign_check = sign_count >= rule_sign_count
                        watch_check = watch_time >= rule_watch_time
                        count_check = watch_count >= rule_watch_count_value
                        is_eligible = sign_check or watch_check or count_check
                        logger.info(f"观众 {viewer_name} 全部条件(OR)规则检查: 签到({sign_count} >= {rule_sign_count} = {sign_check}) 或 观看时长({watch_time} >= {rule_watch_time} = {watch_check}) 或 观看场次({watch_count} >= {rule_watch_count_value} = {count_check}) = {is_eligible}")
                    elif rule_type == RewardRuleType.ALL_AND:
                        sign_check = sign_count >= rule_sign_count
                        watch_check = watch_time >= rule_watch_time
                        count_check = watch_count >= rule_watch_count_value
                        is_eligible = sign_check and watch_check and count_check
                        logger.info(f"观众 {viewer_name} 全部条件(AND)规则检查: 签到({sign_count} >= {rule_sign_count} = {sign_check}) 且 观看时长({watch_time} >= {rule_watch_time} = {watch_check}) 且 观看场次({watch_count} >= {rule_watch_count_value} = {count_check}) = {is_eligible}")
                    
                    logger.info(f"观众 {viewer_name} 最终奖励资格判断: {is_eligible}")
                    
                    # 创建奖励记录
                    reward_record = {
                        'living_id': live_id,
                        'live_viewer_id': viewer_id,
                        'rule_type': rule_type.value,  # 存储枚举的字符串值
                        'rule_sign_count': rule_sign_count,
                        'rule_watch_time': rule_watch_time,
                        'rule_watch_count': rule_watch_count_value,
                        'calculate_batch': batch_id,
                        'reward_amount': rule_amount,  # live_reward_records.reward_amount记录页面设置的金额
                        'is_reward_eligible': is_eligible,
                        'created_at': current_timestamp,
                        'updated_at': current_timestamp
                    }
                    
                    logger.debug(f"创建观众 {viewer_name} 的奖励记录: {reward_record}")
                    
                    # 准备viewer更新数据
                    viewer_update = {
                        'id': viewer_id,
                        'is_reward_eligible': is_eligible,
                        'reward_amount': rule_amount if is_eligible else 0.0,  # live_viewers.reward_amount记录合格后的实际金额
                        'reward_status': '批量计算',
                        'updated_at': current_timestamp
                    }
                    
                    logger.debug(f"准备更新观众 {viewer_name} 的viewer记录: {viewer_update}")
                    
                    reward_records.append(reward_record)
                    viewer_updates.append(viewer_update)
                    
                    # 更新进度
                    with data_lock:
                        processed_viewers += 1
                
                logger.debug(f"完成直播 {live_id} 的一个观众数据块处理，生成了 {len(reward_records)} 条奖励记录")
                return {
                    'reward_records': reward_records,
                    'viewer_updates': viewer_updates
                }
            
            # 创建处理任务 - 每个直播作为一个处理单元
            all_results = []
            processing_tasks = []
            
            update_progress("开始处理数据", 15)
            
            # 创建一个线程池
            with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
                logger.info("创建线程池，开始提交处理任务...")
                
                # 为每个直播创建处理任务
                for live_id, viewers in all_viewers.items():
                    if is_cancelled:
                        break
                    
                    live_config = all_live_configs[live_id]
                    sign_records = all_sign_records.get(live_id, {})
                    
                    # 分批处理观众数据，降低内存使用
                    chunk_size = 200
                    for i in range(0, len(viewers), chunk_size):
                        if is_cancelled:
                            break
                            
                        viewers_chunk = viewers[i:i+chunk_size]
                        logger.info(f"提交任务处理直播 {live_id} 的观众数据，分片大小: {len(viewers_chunk)}，观众数据范围: {i} 到 {i+len(viewers_chunk)-1}")
                        
                        # 提交任务到线程池
                        task = executor.submit(
                            process_chunk, 
                            live_id, 
                            viewers_chunk, 
                            sign_records, 
                            all_watch_counts, 
                            live_config,
                            rule_watch_count
                        )
                        processing_tasks.append(task)
                
                # 等待所有任务完成
                logger.info(f"等待 {len(processing_tasks)} 个处理任务完成...")
                for i, future in enumerate(concurrent.futures.as_completed(processing_tasks)):
                    if update_progress(f"处理中 ({i+1}/{len(processing_tasks)})"):
                        is_cancelled = True
                        logger.warning("用户取消了处理")
                        break
                    
                    # 获取任务结果
                    try:
                        result = future.result()
                        if result:
                            all_results.append(result)
                    except Exception as e:
                        logger.error(f"处理任务发生错误: {str(e)}")
                
            # 如果用户取消了操作
            if is_cancelled:
                logger.warning("处理被用户取消，取消所有后续操作")
                progress_dialog.close()
                QMessageBox.information(self, "提示", "操作已取消。")
                return
            
            # 合并所有结果
            update_progress("正在合并处理结果", 80)
            all_reward_records = []
            all_viewer_updates = []
            
            for result in all_results:
                all_reward_records.extend(result['reward_records'])
                all_viewer_updates.extend(result['viewer_updates'])
            
            logger.info(f"合并所有结果完成，共有 {len(all_reward_records)} 条奖励记录和 {len(all_viewer_updates)} 条viewer更新")
            
            # 批量保存数据到数据库
            update_progress("正在保存数据到数据库", 85)
            success = False
            
            try:
                with self.db_manager.get_session() as session:
                    # 1. 删除旧的奖励记录 - 使用直播ID列表
                    logger.info(f"删除选中的直播ID对应的旧奖励记录: {batch_living_ids}")
                    delete_result = session.query(LiveRewardRecord).filter(
                        LiveRewardRecord.living_id.in_(batch_living_ids)
                    ).delete(synchronize_session=False)
                    logger.info(f"删除旧奖励记录结果: 已删除 {delete_result} 条记录")
                    
                    # 2. 插入新的奖励记录
                    if all_reward_records:
                        logger.info(f"开始批量插入 {len(all_reward_records)} 条新奖励记录...")
                        
                        # 分批插入减轻数据库压力
                        batch_size = 1000
                        for i in range(0, len(all_reward_records), batch_size):
                            batch = all_reward_records[i:i+batch_size]
                            
                            session.execute(
                                LiveRewardRecord.__table__.insert(),
                                [record for record in batch]
                            )
                            
                            logger.info(f"已插入 {i+len(batch)} / {len(all_reward_records)} 条奖励记录")
                    
                    # 3. 更新观众记录
                    if all_viewer_updates:
                        logger.info(f"开始批量更新 {len(all_viewer_updates)} 条观众记录...")
                        
                        # 准备批量更新数据
                        viewer_ids = [update['id'] for update in all_viewer_updates]
                        
                        # 基本信息
                        update_data = {
                            'is_reward_eligible': False,
                            'reward_amount': 0.0,
                            'reward_status': '批量计算',
                            'updated_at': current_timestamp
                        }
                        
                        # 批量更新所有记录
                        session.query(LiveViewer).filter(
                            LiveViewer.id.in_(viewer_ids)
                        ).update(update_data, synchronize_session=False)
                        
                        # 逐个更新符合条件的记录
                        eligible_updates = [u for u in all_viewer_updates if u['is_reward_eligible']]
                        logger.info(f"其中有 {len(eligible_updates)} 条观众记录符合奖励条件，需要单独更新")
                        
                        for update in eligible_updates:
                            session.query(LiveViewer).filter(
                                LiveViewer.id == update['id']
                            ).update({
                                'is_reward_eligible': True,
                                'reward_amount': update['reward_amount']
                            }, synchronize_session=False)
                    
                    # 提交事务
                    logger.info("提交所有数据库操作...")
                    session.commit()
                    success = True
                    logger.info("数据保存成功！")
                    
            except Exception as e:
                logger.error(f"保存数据到数据库时出错: {str(e)}")
                QMessageBox.critical(self, "错误", f"保存数据到数据库时出错: {str(e)}")
            
            # 关闭进度对话框
            progress_dialog.close()
            
            # 显示处理结果
            if success:
                eligible_count = sum(1 for r in all_reward_records if r['is_reward_eligible'])
                total_reward = sum(r['reward_amount'] for r in all_viewer_updates if r['is_reward_eligible'])
                
                logger.info(f"奖励计算完成！处理了 {len(all_reward_records)} 条记录，符合奖励条件: {eligible_count} 人，奖励总金额: ¥{total_reward:.2f}")
                
                QMessageBox.information(
                    self, 
                    "处理完成", 
                    f"奖励计算完成！\n\n"
                    f"处理了 {len(all_reward_records)} 条记录\n"
                    f"符合奖励条件: {eligible_count} 人\n"
                    f"不符合条件: {len(all_reward_records) - eligible_count} 人\n"
                    f"奖励总金额: ¥{total_reward:.2f}\n\n"
                    f"计算批次: {batch_id}"
                )
                
        except Exception as e:
            logger.error(f"处理过程出错: {str(e)}")
            logger.exception("详细错误信息：")
            QMessageBox.critical(self, "错误", f"处理过程出错: {str(e)}")
            progress_dialog.close()
        
        logger.info("==================== 结束计算奖励 ====================")
        
    def composite_export_data(self):
        """导出数据"""
        # 获取或创建logger，避免多次导入
        import logging
        logger = logging.getLogger(__name__)
        
        # 声明进度对话框变量，确保finally块中可以访问
        progress = None
        
        try:
            # 1. 初始检查 - 确保选择了直播
            if not self.selected_live_ids:
                QMessageBox.warning(self, "警告", "请至少选择一场直播")
                return
                
            selected_lives = list(self.selected_live_ids)
            logger.info(f"开始导出数据，选中的直播ID: {selected_lives}")
            
            # 辅助函数：处理用户取消操作
            def handle_cancel(session, message, close_progress=True, rollback=True):
                nonlocal progress
                if close_progress and progress:
                    progress.close()
                logger.info(message)
                if rollback:
                    session.rollback()  # 回滚所有未提交的数据库更改
                return False
            
            # 使用单个数据库会话执行所有操作
            with self.db_manager.get_session() as session:
                from src.models.live_reward_record import LiveRewardRecord
                from src.models.live_sign_record import LiveSignRecord
                from src.models.live_viewer import LiveViewer
                from src.models.living import Living
                from sqlalchemy import func, text, and_
                
                # ===== 校验阶段 =====
                # 创建验证阶段进度对话框
                progress = QProgressDialog("正在验证数据...", "取消", 0, 100, self)
                progress.setWindowTitle("验证数据")
                progress.setMinimumDuration(0)
                progress.setWindowModality(Qt.WindowModal)
                progress.setAutoClose(True)
                progress.setValue(0)
                
                # --- 1. 检查奖励记录 ---
                progress.setLabelText("检查奖励记录...")
                progress.setValue(5)
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在验证开始阶段取消了操作")
                
                # 使用适合单个或多个ID的SQL语法
                if len(selected_lives) > 1:
                    reward_count_result = session.execute(
                        text("SELECT COUNT(*) FROM live_reward_records WHERE living_id IN :ids"),
                        {"ids": tuple(selected_lives)}
                    )
                else:
                    reward_count_result = session.execute(
                        text("SELECT COUNT(*) FROM live_reward_records WHERE living_id = :id"),
                        {"id": selected_lives[0]}
                    )
                reward_count = reward_count_result.scalar()
                
                if reward_count == 0:
                    progress.close()
                    QMessageBox.warning(self, "警告", "请先计算奖励结果后再导出数据")
                    return
                
                # --- 2. 预先获取所有直播信息 ---
                # 一次性获取所有选中直播的信息，避免多次查询
                livings_data = session.query(
                    Living.id, 
                    Living.theme, 
                    Living.livingid
                ).filter(
                    Living.id.in_(selected_lives)
                ).all()
                
                # 创建ID到livingid和theme的映射，用于后续使用
                living_id_to_info = {
                    living.id: {'livingid': living.livingid, 'theme': living.theme} 
                    for living in livings_data
                }
                
                # --- 3. 检查规则一致性 ---
                progress.setLabelText("检查规则一致性...")
                progress.setValue(20)
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在规则检查前取消了操作")
                
                # 获取当前规则设置
                rule_type_value = self.rule_type_combo.currentData()
                logger.info(f"当前界面选择的规则类型: {rule_type_value}, 类型: {type(rule_type_value).__name__}")
                
                # 获取当前页面上的规则设置
                current_rules = {
                    "min_watch_count": self.min_watch_count_spin.value()
                }
                
                for row in range(self.table.rowCount()):
                    # 获取签到次数设置
                    sign_count_spin = self.table.cellWidget(row, 4)
                    if sign_count_spin:
                        current_rules[f"sign_count_{row}"] = sign_count_spin.value()
                    
                    # 获取观看时长设置
                    watch_time_spin = self.table.cellWidget(row, 5)
                    if watch_time_spin:
                        current_rules[f"watch_time_{row}"] = watch_time_spin.value()
                    
                    # 获取红包金额设置
                    reward_spin = self.table.cellWidget(row, 6)
                    if reward_spin:
                        current_rules[f"reward_{row}"] = reward_spin.value()
                
                # 逐个检查每个选中直播的奖励记录
                missing_reward_lives = []  # 存储没有奖励记录的直播ID
                rule_mismatch_lives = []   # 存储规则不匹配的直播ID
                
                # 批量查询所有选中直播的所有规则记录
                try:
                    # 使用text执行原始SQL查询避免枚举转换错误
                    if len(selected_lives) > 1:
                        sql = text("SELECT living_id, rule_type, rule_watch_count FROM live_reward_records WHERE living_id IN :ids")
                        params = {"ids": tuple(selected_lives)}
                    else:
                        sql = text("SELECT living_id, rule_type, rule_watch_count FROM live_reward_records WHERE living_id = :id")
                        params = {"id": selected_lives[0]}
                    
                    all_rewards_result = session.execute(sql, params)
                    
                    # 按直播ID分组规则记录
                    live_id_to_rewards = {}
                    
                    # 创建一个类似的数据结构来保存查询结果
                    RewardRecord = namedtuple('RewardRecord', ['living_id', 'rule_type', 'rule_watch_count'])
                    
                    for row in all_rewards_result:
                        living_id = row[0]
                        rule_type = row[1]  # 这是字符串类型
                        rule_watch_count = row[2]
                        
                        # 创建自定义对象
                        reward = RewardRecord(living_id=living_id, rule_type=rule_type, rule_watch_count=rule_watch_count)
                        
                        logger.info(f"从数据库获取的规则记录: 直播ID={living_id}, 规则类型={rule_type}, 类型={type(rule_type).__name__}, 最小观看场次={rule_watch_count}")
                        
                        if living_id not in live_id_to_rewards:
                            live_id_to_rewards[living_id] = []
                        live_id_to_rewards[living_id].append(reward)
                except Exception as e:
                    logger.error(f"查询规则记录时出错: {str(e)}")
                    progress.close()
                    QMessageBox.critical(self, "错误", f"查询规则记录时出错: {str(e)}")
                    return
                
                # 检查每个直播的规则
                for i, live_id in enumerate(selected_lives):
                    # 更新进度
                    progress.setValue(20 + int(20 * i / len(selected_lives)))
                    
                    if progress.wasCanceled():
                        return handle_cancel(session, f"用户在检查直播 {live_id} 的规则时取消了操作")
                    
                    # 获取该直播的规则记录
                    rewards = live_id_to_rewards.get(live_id, [])
                    
                    if not rewards:
                        missing_reward_lives.append(live_id)
                        continue
                    
                    # 随机取1-2条进行检查
                    import random
                    sample_size = min(2, len(rewards))
                    reward_samples = random.sample(rewards, sample_size)
                    
                    # 从映射中获取直播标题
                    live_title = living_id_to_info.get(live_id, {}).get('theme', f"直播ID: {live_id}")
                    
                    # 检查奖励规则是否与当前设置一致
                    for reward in reward_samples:
                        db_rule_type = reward.rule_type
                        ui_rule_type = rule_type_value
                        
                        logger.info(f"直播 '{live_title}' - 数据库规则类型: {db_rule_type}, UI规则类型: {ui_rule_type}")
                        
                        # 添加详细的规则类型比较调试信息
                        logger.info(f"db_rule_type类型: {type(db_rule_type).__name__}, ui_rule_type类型: {type(ui_rule_type).__name__}")
                        
                        # 使用RewardRuleType.from_string方法将字符串和枚举值进行标准化
                        try:
                            # 将数据库和UI的规则类型转换为枚举对象进行比较
                            db_enum = RewardRuleType.from_string(db_rule_type)
                            ui_enum = RewardRuleType.from_string(ui_rule_type)
                            
                            # 详细记录比较过程
                            logger.info(f"规则类型比较 - DB原始值: '{db_rule_type}' -> 枚举: {db_enum}(值: {db_enum.value}), "
                                       f"UI原始值: '{ui_rule_type}' -> 枚举: {ui_enum}(值: {ui_enum.value})")
                            
                            # 检查规则类型是否一致
                            if db_enum != ui_enum:
                                logger.warning(f"规则类型不匹配! DB枚举: {db_enum} != UI枚举: {ui_enum}")
                                rule_mismatch_lives.append((live_id, live_title, "规则类型不一致"))
                                break
                            else:
                                logger.info(f"规则类型匹配成功: {db_enum} == {ui_enum}")
                        
                        except Exception as e:
                            # 记录转换过程中的任何错误
                            logger.error(f"比较规则类型时出错: {str(e)}")
                            # 回退到简单字符串比较
                            db_str = str(db_rule_type).lower().strip()
                            ui_str = str(ui_rule_type).lower().strip()
                            if db_str != ui_str:
                                logger.warning(f"回退比较: 规则类型不匹配! DB值: '{db_str}' != UI值: '{ui_str}'")
                                rule_mismatch_lives.append((live_id, live_title, "规则类型不一致"))
                                break
                        
                        # 检查最少观看场次是否一致（如果数据库中有这个字段）
                        if hasattr(reward, 'rule_watch_count') and reward.rule_watch_count != current_rules.get("min_watch_count", 0):
                            logger.warning(f"最少观看场次不一致! DB值: {reward.rule_watch_count}, UI值: {current_rules.get('min_watch_count', 0)}")
                            rule_mismatch_lives.append((live_id, live_title, "最少观看场次不一致"))
                            break
                
                progress.setValue(40)
                
                # 处理没有奖励记录的直播
                if missing_reward_lives:
                    missing_live_titles = []
                    for live_id in missing_reward_lives:
                        title = living_id_to_info.get(live_id, {}).get('theme', f"直播ID: {live_id}")
                        missing_live_titles.append(title)
                    
                    progress.close()
                    result = QMessageBox.warning(
                        self,
                        "缺少奖励数据",
                        f"以下直播尚未计算奖励数据，请先计算后再导出：\n\n{', '.join(missing_live_titles)}",
                        QMessageBox.Ok
                    )
                    return
                
                # 处理规则不匹配的直播
                if rule_mismatch_lives:
                    # 格式化不匹配信息
                    mismatch_info = "\n".join([f"• {title} ({reason})" for _, title, reason in rule_mismatch_lives])
                    
                    progress.close()
                    result = QMessageBox.question(
                        self,
                        "规则不一致",
                        f"以下直播的奖励规则与当前导出规则不一致：\n\n{mismatch_info}\n\n是否重新计算奖励后再导出？",
                        QMessageBox.Yes | QMessageBox.No,
                        QMessageBox.Yes
                    )
                    
                    if result == QMessageBox.Yes:
                        # 用户选择重新计算
                        self.calculate_reward()
                        return
                    else:
                        # 用户选择继续导出，重新显示进度对话框
                        progress = QProgressDialog("正在验证数据...", "取消", 0, 100, self)
                        progress.setWindowTitle("验证数据")
                        progress.setMinimumDuration(0)
                        progress.setWindowModality(Qt.WindowModal)
                        progress.setAutoClose(True)
                        progress.setValue(40)
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在规则不一致处理后取消了操作")
                
                # --- 4. 检查签到记录 ---
                progress.setLabelText("检查签到记录...")
                progress.setValue(60)
                
                # 批量查询签到记录数
                missing_sign_records = []
                
                # 使用一次性查询获取所有直播的签到记录数
                sign_records_query = session.query(
                    Living.id,
                    func.count(LiveSignRecord.id).label('sign_count')
                ).outerjoin(
                    LiveSignRecord,
                    LiveSignRecord.living_id == Living.livingid
                ).filter(
                    Living.id.in_(selected_lives)
                ).group_by(
                    Living.id
                ).all()
                
                # 检查每个直播的签到记录
                for record in sign_records_query:
                    if record.sign_count == 0:
                        live_title = living_id_to_info.get(record.id, {}).get('theme', f"直播ID: {record.id}")
                        missing_sign_records.append(live_title)
                
                if missing_sign_records:
                    progress.close()
                    missing_text = "\n- ".join(missing_sign_records)
                    QMessageBox.warning(
                        self, 
                        "警告", 
                        f"以下直播没有签到记录，请先导入签到后再计算奖励结果:\n- {missing_text}"
                    )
                    return
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在检查签到记录后取消了操作")
                
                # --- 5. 检查观众信息 ---
                progress.setLabelText("检查观众信息...")
                progress.setValue(80)
                
                # 批量查询观众记录数
                missing_viewers = []
                
                # 使用一次性查询获取所有直播的观众记录数
                viewer_records_query = session.query(
                    Living.id,
                    func.count(LiveViewer.id).label('viewer_count')
                ).outerjoin(
                    LiveViewer,
                    LiveViewer.living_id == Living.id
                ).filter(
                    Living.id.in_(selected_lives)
                ).group_by(
                    Living.id
                ).all()
                
                # 检查每个直播的观众记录
                for record in viewer_records_query:
                    if record.viewer_count == 0:
                        live_title = living_id_to_info.get(record.id, {}).get('theme', f"直播ID: {record.id}")
                        missing_viewers.append(live_title)
                
                if missing_viewers:
                    progress.close()
                    missing_text = "\n- ".join(missing_viewers)
                    QMessageBox.warning(
                        self, 
                        "警告", 
                        f"以下直播没有观众信息，请先拉取观众信息后再计算奖励结果:\n- {missing_text}"
                    )
                    return
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在检查观众信息后取消了操作")
                
                progress.setValue(100)
                progress.close()
                
                # ===== 数据导出阶段 =====
                # 完成校验后，再选择文件保存位置
                from datetime import datetime
                current_datetime = datetime.now().strftime("%Y%m%d%H%M%S")
                default_filename = f"综合导出直播数据_{current_datetime}.xlsx"
                
                # 选择导出文件
                file_path, _ = QFileDialog.getSaveFileName(
                    self,
                    "导出奖励结果",
                    default_filename,
                    "Excel Files (*.xlsx)"
                )
                
                if not file_path:
                    logger.info("用户取消了文件选择")
                    return  # 用户取消了选择
                
                # 显示数据处理进度对话框
                progress = QProgressDialog("正在导出数据...", "取消", 0, 100, self)
                progress.setWindowTitle("导出数据")
                progress.setMinimumDuration(0)
                progress.setWindowModality(Qt.WindowModal)
                progress.setAutoClose(True)
                progress.setValue(0)
                
                # 预先准备数据导出
                import pandas as pd
                from openpyxl import Workbook
                from openpyxl.styles import PatternFill
                from openpyxl.utils.dataframe import dataframe_to_rows
                from openpyxl.utils import get_column_letter
                
                # 获取用户记录总数以便显示进度
                total_viewers_count = session.query(func.count(func.distinct(LiveViewer.userid))).filter(
                    LiveViewer.living_id.in_(selected_lives)
                ).scalar() or 0
                
                logger.info(f"预计需要处理 {total_viewers_count} 条用户记录")
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在数据准备阶段取消了操作")
                
                # 创建Excel工作簿
                wb = Workbook()
                ws = wb.active
                ws.title = "综合直播数据"
                
                # 设置标题行
                headers = ["用户ID", "用户名称"]
                
                # 为每个直播添加三列数据
                for i, live_id in enumerate(selected_lives, 1):
                    title = living_id_to_info[live_id]['theme']
                    headers.extend([
                        f"观看时长{i} ({title})",
                        f"签到次数{i} ({title})",
                        f"邀请人{i} ({title})"
                    ])
                
                # 添加合计和资格列
                headers.extend(["红包合计($)", "是否合格"])
                
                # 写入标题行
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
                
                # 为不同直播的列添加不同颜色
                colors = ['FFCCCC', 'CCFFCC', 'CCCCFF', 'FFFFCC', 'FFCCFF', 'CCFFFF']
                
                # 设置标题行颜色
                for i, live_id in enumerate(selected_lives, 1):
                    color = colors[i % len(colors)]  # 循环使用颜色
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    # 为每个直播的3列数据设置颜色
                    col_start = 3 + (i-1)*3  # 从第3列开始是第一个直播的数据
                    for col in range(col_start, col_start+3):
                        ws.cell(row=1, column=col).fill = fill
                
                progress.setValue(10)
                
                if progress.wasCanceled():
                    return handle_cancel(session, "用户在表格准备阶段取消了操作")
                
                # 批量处理用户数据
                # 获取所有用户ID - 使用yield_per优化内存使用
                all_viewer_ids_query = session.query(LiveViewer.userid).distinct().filter(
                    LiveViewer.living_id.in_(selected_lives)
                )
                
                # 当前行索引，从第2行开始（第1行是标题）
                current_row = 2
                
                # 批处理大小
                batch_size = 1000
                
                # 记录所有已处理的用户ID（用于更新reward_status）
                processed_user_ids = set()
                
                # 获取总批次数
                total_batches = (total_viewers_count + batch_size - 1) // batch_size
                
                logger.info(f"总共需要处理 {total_batches} 批数据，每批 {batch_size} 条")
                
                # 是否已有数据写入Excel
                has_data_written = False
                
                try:
                    # 辅助函数：转换秒为时分秒格式
                    def format_seconds(seconds):
                        if not seconds:
                            return "未观看"
                        hours, remainder = divmod(seconds, 3600)
                        minutes, seconds = divmod(remainder, 60)
                        return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    
                    # 分批处理用户数据
                    for batch_index in range(total_batches):
                        if progress.wasCanceled():
                            logger.info("用户在数据处理阶段取消了操作")
                            raise CancelOperationException("用户取消了操作")
                        
                        # 计算当前批次的进度范围：10% - 80%，总共70%的进度用于处理数据
                        progress_start = 10
                        progress_end = 80
                        progress_per_batch = (progress_end - progress_start) / total_batches
                        current_progress = progress_start + batch_index * progress_per_batch
                        progress.setValue(int(current_progress))
                        progress.setLabelText(f"正在处理批次 {batch_index + 1}/{total_batches}...")
                        
                        # 获取当前批次的用户ID
                        viewer_ids_batch = all_viewer_ids_query.limit(batch_size).offset(batch_index * batch_size).all()
                        
                        if not viewer_ids_batch:
                            logger.info(f"批次 {batch_index + 1} 没有数据，跳过")
                            continue
                        
                        logger.info(f"处理批次 {batch_index + 1}/{total_batches}，包含 {len(viewer_ids_batch)} 条用户ID")
                        
                        # 提取用户ID列表
                        user_ids = [uid for (uid,) in viewer_ids_batch]
                        processed_user_ids.update(user_ids)
                        
                        # 批量查询用户基本信息
                        basic_info_query = session.query(
                            LiveViewer.userid,
                            LiveViewer.name,
                            func.sum(LiveViewer.reward_amount).label('total_reward'),
                            func.max(LiveViewer.is_reward_eligible).label('is_eligible')
                        ).filter(
                            LiveViewer.userid.in_(user_ids),
                            LiveViewer.living_id.in_(selected_lives)
                        ).group_by(
                            LiveViewer.userid,
                            LiveViewer.name
                        ).all()
                        
                        # 创建用户ID到基本信息的映射
                        user_id_to_basic_info = {
                            info.userid: {
                                'name': info.name,
                                'total_reward': info.total_reward or 0,
                                'is_eligible': info.is_eligible or False
                            }
                            for info in basic_info_query
                        }
                        
                        # 批量查询用户观看记录 - 使用一次性查询优化
                        viewer_info_dict = {}
                        
                        # 一次性查询所有选中直播的所有用户观看记录
                        all_viewer_info = session.query(
                            LiveViewer.userid,
                            LiveViewer.living_id,
                            LiveViewer.watch_time,
                            LiveViewer.sign_count,
                            LiveViewer.invitor_name
                        ).filter(
                            LiveViewer.userid.in_(user_ids),
                            LiveViewer.living_id.in_(selected_lives)
                        ).all()
                        
                        # 构建查询结果字典
                        for info in all_viewer_info:
                            viewer_info_dict[(info.userid, info.living_id)] = {
                                'watch_time': info.watch_time,
                                'sign_count': info.sign_count,
                                'invitor_name': info.invitor_name
                            }
                        
                        # 为每个用户生成一行数据
                        for user_id in user_ids:
                            if user_id not in user_id_to_basic_info:
                                continue  # 跳过没有基本信息的用户
                            
                            has_data_written = True  # 标记已有数据写入
                            
                            basic_info = user_id_to_basic_info[user_id]
                            
                            # 在当前行写入基本信息
                            ws.cell(row=current_row, column=1, value=user_id)
                            ws.cell(row=current_row, column=2, value=basic_info['name'])
                            
                            # 为每个直播添加详细信息
                            for i, live_id in enumerate(selected_lives, 1):
                                viewer_info = viewer_info_dict.get((user_id, live_id), None)
                                
                                # 计算列索引
                                watch_time_col = 3 + (i-1)*3
                                sign_count_col = watch_time_col + 1
                                invitor_col = watch_time_col + 2
                                
                                # 将秒转换为时分秒格式
                                watch_time_str = format_seconds(viewer_info['watch_time'] if viewer_info else None)
                                
                                # 写入单元格
                                ws.cell(row=current_row, column=watch_time_col, value=watch_time_str)
                                ws.cell(row=current_row, column=sign_count_col, value=viewer_info['sign_count'] if viewer_info else 0)
                                ws.cell(row=current_row, column=invitor_col, value=viewer_info['invitor_name'] if viewer_info and viewer_info['invitor_name'] else "无")
                                
                                # 设置颜色
                                color = colors[i % len(colors)]
                                fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                                for col in range(watch_time_col, invitor_col + 1):
                                    ws.cell(row=current_row, column=col).fill = fill
                            
                            # 添加合计和资格信息
                            last_col = 3 + len(selected_lives)*3
                            ws.cell(row=current_row, column=last_col, value=basic_info['total_reward'])
                            ws.cell(row=current_row, column=last_col+1, value="是" if basic_info['is_eligible'] else "否")
                            
                            # 移动到下一行
                            current_row += 1
                        
                        # 适当控制内存清理
                        if batch_index % 5 == 0:
                            import gc
                            gc.collect()
                    
                    progress.setValue(80)
                    progress.setLabelText("正在保存Excel文件...")
                    
                    # 自动调整列宽
                    for col in range(1, len(headers) + 1):
                        max_length = 0
                        column_letter = get_column_letter(col)
                        # 仅检查前100行和标题行，避免过多计算
                        max_rows = min(current_row - 1, 100)
                        for row in range(1, max_rows + 1):
                            cell_value = str(ws.cell(row=row, column=col).value or "")
                            max_length = max(max_length, len(cell_value))
                        # 设置列宽（字符宽度 + 2的边距）
                        adjusted_width = max_length + 2
                        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)  # 最大宽度限制为50
                    
                    # 保存Excel文件
                    wb.save(file_path)
                    
                    if progress.wasCanceled():
                        return handle_cancel(session, "用户在保存Excel文件后取消了操作", True, True)
                    
                    progress.setValue(85)
                    progress.setLabelText("正在更新数据库状态...")
                    
                    # 批量更新reward_status
                    logger.info(f"更新 {len(processed_user_ids)} 个用户的reward_status")
                    
                    # 准备批量更新语句
                    now = datetime.now()
                    
                    # 分批更新数据库记录，每批最多1000条
                    updated_count = 0
                    batch_size = 1000
                    
                    # 将用户ID列表分成多个批次
                    user_id_batches = [list(processed_user_ids)[i:i+batch_size] for i in range(0, len(processed_user_ids), batch_size)]
                    
                    for batch_index, user_id_batch in enumerate(user_id_batches):
                        if progress.wasCanceled():
                            return handle_cancel(session, f"用户在更新数据库批次 {batch_index + 1}/{len(user_id_batches)} 时取消了操作", True, True)
                        
                        # 当前批次的进度：85% - 95%，总共10%
                        progress_start = 85
                        progress_end = 95
                        progress_per_batch = (progress_end - progress_start) / len(user_id_batches)
                        current_progress = progress_start + batch_index * progress_per_batch
                        progress.setValue(int(current_progress))
                        progress.setLabelText(f"正在更新数据库 {batch_index + 1}/{len(user_id_batches)}...")
                        
                        # 更新当前批次的用户记录
                        viewers_to_update = session.query(LiveViewer).filter(
                            LiveViewer.userid.in_(user_id_batch),
                            LiveViewer.living_id.in_(selected_lives)
                        ).all()
                        
                        for viewer in viewers_to_update:
                            # 追加"已导出"标记（如果不存在）
                            if not viewer.reward_status:
                                viewer.reward_status = "已导出"
                            elif "已导出" not in viewer.reward_status:
                                viewer.reward_status += "，已导出"
                            
                            # 更新时间戳
                            viewer.updated_at = now
                            updated_count += 1
                        
                        # 每批次提交一次，减少内存占用
                        session.commit()
                    
                    logger.info(f"成功更新了 {updated_count} 条记录的reward_status")
                    
                    progress.setValue(98)
                    
                    # 最终提交所有更改
                    session.commit()
                    
                    progress.setValue(100)
                    progress.close()
                    
                    # 显示成功消息
                    # QMessageBox.information(
                    #     self, 
                    #     "成功", 
                    #     f"已成功导出 {current_row - 2} 条用户记录到文件:\n{file_path}\n\n"
                    #     f"已更新 {updated_count} 条记录的状态。"
                    # )

                    progress.setValue(100)
                    # 进度对话框在这里关闭一次就足够了
                    progress.close()
                    
                    # 创建一个自定义的消息框，包含打开文件按钮
                    msg_box = QMessageBox(self)
                    msg_box.setWindowTitle("导出成功")
                    msg_box.setIcon(QMessageBox.Information)
                    msg_box.setText(f"已成功导出 {current_row - 2} 条用户记录！")
                    msg_box.setInformativeText(f"文件保存在：\n{file_path}\n\n已更新 {updated_count} 条记录的状态。")
                    
                    # 添加按钮
                    open_button = msg_box.addButton("打开文件", QMessageBox.ActionRole)
                    close_button = msg_box.addButton("关闭", QMessageBox.RejectRole)
                    
                    # 显示消息框，等待用户操作
                    msg_box.exec()
                    
                    # 处理用户的按钮点击
                    clicked_button = msg_box.clickedButton()
                    if clicked_button == open_button:
                        # 使用系统默认应用程序打开文件
                        import subprocess
                        import platform
                        
                        try:
                            system = platform.system()
                            if system == 'Darwin':  # macOS
                                subprocess.call(('open', file_path))
                            elif system == 'Windows':
                                os.startfile(file_path)
                            else:  # Linux
                                subprocess.call(('xdg-open', file_path))
                            logger.info(f"用户选择打开导出文件: {file_path}")
                        except Exception as e:
                            logger.error(f"打开文件失败: {str(e)}")
                            QMessageBox.warning(self, "警告", f"无法打开文件，请手动查找：\n{file_path}")
                    else:
                        logger.info("用户关闭了导出成功提示框")
                
                except CancelOperationException as e:
                    # 用户取消操作的异常处理
                    return handle_cancel(session, "用户取消了操作", True, True)
            
        except Exception as e:
            logger.error(f"导出数据失败: {str(e)}")
            logger.exception("详细错误信息：")
            QMessageBox.critical(self, "错误", f"导出数据失败: {str(e)}")
            
        finally:
            # 确保进度对话框被关闭
            if progress:
                progress.close()

# 添加自定义异常类，用于处理用户取消操作
class CancelOperationException(Exception):
    """用户取消操作的自定义异常"""
    pass